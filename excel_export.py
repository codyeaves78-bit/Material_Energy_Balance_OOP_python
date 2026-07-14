# excel_export: shared openpyxl helpers so every unit operation exports to
# the same workbook with one consistent look.
#
# Usage pattern (mirrors neat_display):
#     from excel_export import new_workbook
#     wb = new_workbook()
#     mill_floor.to_excel(wb)      # each class writes its own sheet
#     clarification.to_excel(wb)
#     wb.save("factory_balance.xlsx")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break


# ── Palette ────────────────────────────────────────────────────────────────
NAVY        = "1F4E78"   # title bar
STEEL       = "305496"   # table header rows
LIGHT_BLUE  = "DDEBF7"   # section header fill
BAND_GREY   = "F2F2F2"   # alternating table rows
BORDER_GREY = "BFBFBF"

TITLE_FONT    = Font(bold=True, size=13, color="FFFFFF")
SUBTITLE_FONT = Font(italic=True, size=10, color="DDEBF7")
SECTION_FONT  = Font(bold=True, size=11, color=NAVY)
HEADER_FONT   = Font(bold=True, size=10, color="FFFFFF")
LABEL_FONT    = Font(size=10)
VALUE_FONT    = Font(size=10)
UNIT_FONT     = Font(italic=True, size=9, color="808080")
TOTAL_FONT    = Font(bold=True, size=10)

TITLE_FILL   = PatternFill("solid", start_color=NAVY)
SECTION_FILL = PatternFill("solid", start_color=LIGHT_BLUE)
HEADER_FILL  = PatternFill("solid", start_color=STEEL)
BAND_FILL    = PatternFill("solid", start_color=BAND_GREY)

_thin    = Side(style="thin", color=BORDER_GREY)
BOX      = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
TOTALTOP = Border(top=Side(style="medium", color=NAVY))

LEFT   = Alignment(horizontal="left",   vertical="center")
RIGHT  = Alignment(horizontal="right",  vertical="center")
CENTER = Alignment(horizontal="center", vertical="center")


def _display_text(value, fmt):
    """Text as it will actually be displayed in the cell, so column
    auto-sizing matches the real rendered width (not an assumed .2f)."""
    if not isinstance(value, (int, float)):
        return str(value or "")
    if not fmt or fmt == "@":
        return f"{value:,.2f}"
    decimals = fmt.split(".")[1].count("0") if "." in fmt else 0
    if "%" in fmt:
        return f"{value * 100:,.{decimals}f}%"
    return f"{value:,.{decimals}f}"


def new_workbook() -> Workbook:
    """Fresh workbook with the default empty sheet removed."""
    wb = Workbook()
    wb.remove(wb.active)
    return wb


class SheetWriter:
    """
    Writes one styled worksheet top-to-bottom, the same way neat_display
    prints: a title bar, section headers, label/value/unit rows, and tables.

    Parameters
    ----------
    workbook   : openpyxl Workbook to add the sheet to.
    sheet_name : Tab name (trimmed to Excel's 31-char limit, deduped).
    ncols      : Width of the sheet in columns — title/section bars span this.
    """

    def __init__(self, workbook: Workbook, sheet_name: str, ncols: int = 7):
        name = sheet_name[:31]
        i = 2
        while name in workbook.sheetnames:
            suffix = f" ({i})"
            name = sheet_name[:31 - len(suffix)] + suffix
            i += 1
        self.ws = workbook.create_sheet(name)
        self.ncols = ncols
        self.r = 1                              # current row
        self._widths = {}                       # col index -> max content width
        self.ws.sheet_view.showGridLines = False

        # Print setup: portrait letter paper, scaled to fit one page wide.
        # Reports still span multiple pages down for tall sheets, just never
        # sideways — use page_break() to control where those page breaks fall.
        ws = self.ws
        ws.page_setup.orientation = 'portrait'
        ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_margins.left = ws.page_margins.right = 0.4
        ws.page_margins.top = ws.page_margins.bottom = 0.5
        ws.page_margins.header = ws.page_margins.footer = 0.2

    # ── internals ──────────────────────────────────────────────────────────

    def _cell(self, col, value, font=VALUE_FONT, fill=None, align=None,
              fmt=None, border=None, track_width=True):
        c = self.ws.cell(row=self.r, column=col, value=value)
        c.font = font
        if fill:
            c.fill = fill
        if align:
            c.alignment = align
        if border:
            c.border = border
        if fmt and isinstance(value, (int, float)):
            c.number_format = fmt
        # track content width for auto-sizing, using the actual displayed
        # text (decimals as shown by fmt) so estimates match real width
        if track_width:
            self._widths[col] = max(self._widths.get(col, 0), len(_display_text(value, fmt)))
        return c

    def _merge_bar(self, text, font, fill):
        self.ws.merge_cells(start_row=self.r, start_column=1,
                            end_row=self.r, end_column=self.ncols)
        # text spans the whole merged bar, not just column A — don't let it
        # dictate column A's auto-fit width
        self._cell(1, text, font=font, fill=fill, align=LEFT, track_width=False)
        for col in range(2, self.ncols + 1):
            self.ws.cell(row=self.r, column=col).fill = fill
        self.r += 1

    # ── building blocks ────────────────────────────────────────────────────

    def title(self, text, subtitle=None):
        self.ws.row_dimensions[self.r].height = 22
        self._merge_bar(text, TITLE_FONT, TITLE_FILL)
        if subtitle:
            self.ws.row_dimensions[self.r].height = 16
            self._merge_bar(subtitle, SUBTITLE_FONT, TITLE_FILL)
        self.blank()
        self.ws.freeze_panes = self.ws.cell(row=self.r, column=1)

    def section(self, text):
        self.blank()
        self.ws.row_dimensions[self.r].height = 17
        self._merge_bar(text, SECTION_FONT, SECTION_FILL)

    def row(self, label, value, unit="", fmt="#,##0.00", col=1):
        """One 'label  value  unit' line, like neat_display's row(). Pass
        col to start the label at a different column (e.g. to move a long
        label out from under a narrow column A)."""
        self._cell(col, label, font=LABEL_FONT, align=LEFT)
        value_align = RIGHT if isinstance(value, (int, float)) else LEFT
        self._cell(col + 1, value, font=VALUE_FONT, align=value_align, fmt=fmt)
        if unit:
            self._cell(col + 2, unit, font=UNIT_FONT, align=LEFT)
        self.r += 1

    def table(self, headers, rows, fmts=None, totals=None):
        """
        Bordered table with a colored header row and banded body rows.

        headers : list of column titles
        rows    : list of row tuples (str or numeric cells)
        fmts    : per-column number formats (defaults to #,##0.00)
        totals  : optional list of row tuples rendered bold above a rule
        """
        n = len(headers)
        fmts = fmts or ["#,##0.00"] * n

        for j, h in enumerate(headers, start=1):
            self._cell(j, h, font=HEADER_FONT, fill=HEADER_FILL,
                       align=CENTER, border=BOX)
        self.r += 1

        for i, row in enumerate(rows):
            band = BAND_FILL if i % 2 else None
            for j, v in enumerate(row, start=1):
                align = RIGHT if isinstance(v, (int, float)) else LEFT
                self._cell(j, v, font=VALUE_FONT, fill=band, align=align,
                           fmt=fmts[j - 1], border=BOX)
            self.r += 1

        for row in totals or []:
            for j in range(1, n + 1):
                v = row[j - 1] if j - 1 < len(row) else None
                align = RIGHT if isinstance(v, (int, float)) else LEFT
                self._cell(j, v, font=TOTAL_FONT, align=align,
                           fmt=fmts[j - 1], border=TOTALTOP)
            self.r += 1

    def row_pair(self, left, right, gap_col=5):
        """Write two independent (label, value, unit, fmt) row-lists side by
        side on the same rows — e.g. PARAMETERS next to KEY RESULTS — so the
        block is half as tall as writing them one after another."""
        for i in range(max(len(left), len(right))):
            if i < len(left):
                label, value, unit, fmt = left[i]
                self._cell(1, label, font=LABEL_FONT, align=LEFT)
                self._cell(2, value, font=VALUE_FONT, align=RIGHT, fmt=fmt)
                if unit:
                    self._cell(3, unit, font=UNIT_FONT, align=LEFT)
            if i < len(right):
                label, value, unit, fmt = right[i]
                self._cell(gap_col, label, font=LABEL_FONT, align=LEFT)
                self._cell(gap_col + 1, value, font=VALUE_FONT, align=RIGHT, fmt=fmt)
                if unit:
                    self._cell(gap_col + 2, unit, font=UNIT_FONT, align=LEFT)
            self.r += 1

    def blank(self, n=1):
        self.r += n

    def page_break(self):
        """Force whatever is written next to start at the top of a new
        printed page (e.g. between units sharing one sheet)."""
        if self.r > 1:
            self.ws.row_breaks.append(Break(id=self.r - 1))

    def image(self, fig, anchor_col=1, scale=1.0, width_in=None):
        """Embed a matplotlib Figure at the current row. Pass width_in to
        size it to an absolute width in inches (Excel renders images at
        96 DPI); otherwise scale multiplies the native saved pixel size."""
        import io
        from openpyxl.drawing.image import Image as XLImage
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=150, bbox_inches="tight")
        buf.seek(0)
        img = XLImage(buf)
        if width_in is not None:
            scale = (width_in * 96) / img.width
        img.width  *= scale
        img.height *= scale
        anchor = f"{get_column_letter(anchor_col)}{self.r}"
        self.ws.add_image(img, anchor)
        # advance past the image (~18 px per row)
        self.r += int(img.height / 18) + 2

    def finish(self, min_width=4, pad=2):
        """Auto-size columns from the widest content written to each — like
        double-clicking a column border's right edge in Excel."""
        for col, w in self._widths.items():
            letter = get_column_letter(col)
            self.ws.column_dimensions[letter].width = max(min_width, w + pad)
        return self.ws

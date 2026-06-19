# Boiling scheme functions for sugar mill material and energy balance.
# Each function chains Pan and Centrifugal objects, iterates until converged,
# and returns all objects plus stream summaries in a flat result dict.

from SugarStream import SugarStream
from Pan import Pan
from Centrifugal import Centrifugal


def three_boiling_double_magma(
    syrup,
    a_pan_specs, b_pan_specs, c_pan_specs, seed_pan_specs,
    a_cent_specs, b_cent_specs, c_cent_specs,
    b_magma_brix=92.0,
    c_magma_brix=92.0,
    b_magma_temp_F=140.0,
    c_magma_temp_F=140.0,
    a_mol_to_seed_pct=0.10,
    b_mol_to_seed_pct=0.10,
    b_remelt_pct=25.0,
    c_remelt_pct=25.0,
    remelt_brix=65.0,
    max_iter=60,
    tol=0.1,
):
    """
    Three-massecuite double magma boiling scheme.

    ALL massecuite at each grade goes to its centrifugal — no massecuite is split.
    "Magma" is sugar crystals from a centrifugal mixed with water and recycled as footing.

    Cascade:
        syrup + B magma                  →  A pan  →  A massecuite  →  A centrifugal
            A centrifugal                →  A sugar (warehouse) + A molasses
        A molasses (1-a%) + C magma      →  B pan  →  B massecuite  →  B centrifugal
            B centrifugal                →  B sugar + B molasses
            B sugar + water              →  B magma  (recycled as A pan footing)
        A molasses (a%) + B molasses (b%)→  Seed pan  →  Seed massecuite  (no centrifugal)
        B molasses (1-b%) + Seed masse   →  C pan  →  C massecuite  →  C centrifugal
            C centrifugal                →  C sugar + Final molasses
            C sugar + water              →  C magma  (recycled as B pan footing)

    syrup               : SugarStream — clarified juice/syrup entering A pan

    *_pan_specs  (dict) : kwargs for Pan(), excluding feed_streams:
                          heating_surface_ft2, inches_vacuum, supersaturation,
                          head_ft, masse_brix, ml_purity, steam_type, heat_loss_factor

    seed_pan_specs      : same as *_pan_specs — its massecuite goes entirely to C pan,
                          no centrifugal. Set ml_purity < expected feed purity (~60-65%).

    *_cent_specs (dict) : kwargs for Centrifugal(), excluding massecuite and flow:
                          target_molasses_brix, purity_rise, sugar_purity, sugar_moisture

    b_magma_brix        : target Brix when B sugar is dissolved in water [default 92]
    c_magma_brix        : target Brix when C sugar is dissolved in water [default 92]
    b_magma_temp_F      : temperature of the B magma stream entering A pan [default 120°F]
    c_magma_temp_F      : temperature of the C magma stream entering B pan [default 120°F]
    a_mol_to_seed_pct   : fraction of A molasses sent to Seed pan [default 0.10]
    b_mol_to_seed_pct   : fraction of B molasses sent to Seed pan [default 0.10]
    b_remelt_pct        : % of total B magma sent to remelt instead of A pan footing [default 25]
    c_remelt_pct        : % of total C magma sent to remelt instead of B pan footing [default 25]
    remelt_brix         : Brix at which remelt streams are dissolved before entering A pan [default 65]
    """
    if a_mol_to_seed_pct <= 0 and b_mol_to_seed_pct <= 0:
        raise ValueError(
            "At least one of a_mol_to_seed_pct or b_mol_to_seed_pct must be > 0."
        )

    # Start with zero-flow magma and seed streams — Pan handles zero-flow feeds gracefully
    b_magma_to_a = SugarStream(
        brix=b_magma_brix, purity=b_cent_specs['sugar_purity'],
        flow_lb_per_hr=0.0, temp_deg_F=b_magma_temp_F,
        pressure_psia=14.7, level_ft=0,
    )
    b_remelt = SugarStream(
        brix=remelt_brix, purity=b_cent_specs['sugar_purity'],
        flow_lb_per_hr=0.0, temp_deg_F=b_magma_temp_F,
        pressure_psia=14.7, level_ft=0,
    )
    c_magma_to_b = SugarStream(
        brix=c_magma_brix, purity=c_cent_specs['sugar_purity'],
        flow_lb_per_hr=0.0, temp_deg_F=c_magma_temp_F,
        pressure_psia=14.7, level_ft=0,
    )
    c_remelt = SugarStream(
        brix=remelt_brix, purity=c_cent_specs['sugar_purity'],
        flow_lb_per_hr=0.0, temp_deg_F=c_magma_temp_F,
        pressure_psia=14.7, level_ft=0,
    )
    seed_stream = SugarStream(
        brix=seed_pan_specs['masse_brix'], purity=seed_pan_specs['ml_purity'],
        flow_lb_per_hr=0.0, temp_deg_F=150.0, # make temp a user input
        pressure_psia=14.7, level_ft=0,
    )

    converged = False
    a_pan = b_pan = c_pan = seed_pan = None
    a_cent = b_cent = c_cent = None

    for iteration in range(max_iter):
        b_prev    = b_magma_to_a.flow_lb_per_hr
        c_prev    = c_magma_to_b.flow_lb_per_hr
        seed_prev = seed_stream.flow_lb_per_hr

        # --- A pan: syrup + B magma footing + B remelt + C remelt ---
        a_feeds = [syrup]
        if b_magma_to_a.flow_lb_per_hr > 0: a_feeds.append(b_magma_to_a)
        if b_remelt.flow_lb_per_hr > 0:     a_feeds.append(b_remelt)
        if c_remelt.flow_lb_per_hr > 0:     a_feeds.append(c_remelt)
        a_pan  = Pan(feed_streams=a_feeds, **a_pan_specs)
        a_cent = Centrifugal(
            massecuite=a_pan.massecuite,
            massecuite_flow_lb_hr=a_pan.massecuite_flow_lb_hr,
            **a_cent_specs,
        )

        # Split A molasses: fraction to Seed pan, rest to B pan
        a_mol         = a_cent.molasses_stream
        a_mol_to_b    = SugarStream.copy(a_mol, flow_lb_per_hr=a_mol.flow_lb_per_hr * (1.0 - a_mol_to_seed_pct))
        a_mol_to_seed = SugarStream.copy(a_mol, flow_lb_per_hr=a_mol.flow_lb_per_hr * a_mol_to_seed_pct)

        # --- B pan: A molasses (partial) + C magma footing ---
        b_feeds = [a_mol_to_b, c_magma_to_b] if c_magma_to_b.flow_lb_per_hr > 0 else [a_mol_to_b]
        b_pan  = Pan(feed_streams=b_feeds, **b_pan_specs)
        b_cent = Centrifugal(
            massecuite=b_pan.massecuite,
            massecuite_flow_lb_hr=b_pan.massecuite_flow_lb_hr,
            **b_cent_specs,
        )

        # Split B molasses: fraction to Seed pan, rest to C pan
        b_mol         = b_cent.molasses_stream
        b_mol_to_c    = SugarStream.copy(b_mol, flow_lb_per_hr=b_mol.flow_lb_per_hr * (1.0 - b_mol_to_seed_pct))
        b_mol_to_seed = SugarStream.copy(b_mol, flow_lb_per_hr=b_mol.flow_lb_per_hr * b_mol_to_seed_pct)

        # --- Seed pan: A molasses fraction + B molasses fraction ---
        seed_pan = Pan(feed_streams=[a_mol_to_seed, b_mol_to_seed], **seed_pan_specs)
        # ALL seed massecuite goes to C pan as footing — no centrifugal
        seed_stream = SugarStream(
            brix=seed_pan.masse_brix,
            purity=seed_pan.massecuite.masse_purity,
            flow_lb_per_hr=seed_pan.massecuite_flow_lb_hr,
            temp_deg_F=seed_pan.massecuite.massecuite_temp,
            pressure_psia=14.7, level_ft=0,
        )

        # --- C pan: B molasses (partial) + Seed massecuite footing ---
        c_feeds = [b_mol_to_c, seed_stream] if seed_stream.flow_lb_per_hr > 0 else [b_mol_to_c]
        c_pan  = Pan(feed_streams=c_feeds, **c_pan_specs)
        c_cent = Centrifugal(
            massecuite=c_pan.massecuite,
            massecuite_flow_lb_hr=c_pan.massecuite_flow_lb_hr,
            **c_cent_specs,
        )

        # --- Update magma streams (compute total, then split into footing vs. remelt) ---
        # Footing stays at magma_brix; remelt fraction re-dissolved to remelt_brix.
        # AP (pol/solids) is conserved by water addition in both paths.
        b_magma_total_flow = b_cent.crystals_to_sugar_lb_hr / (b_magma_brix / 100.0)
        b_to_a_flow        = b_magma_total_flow * (1.0 - b_remelt_pct / 100.0)
        b_remelt_flow      = (b_magma_total_flow * (b_remelt_pct / 100.0)
                              * (b_magma_brix / 100.0) / (remelt_brix / 100.0))
        b_magma_to_a = SugarStream(
            brix=b_magma_brix, purity=b_cent.sugar_purity,
            flow_lb_per_hr=b_to_a_flow,
            temp_deg_F=b_magma_temp_F, pressure_psia=14.7, level_ft=0,
        )
        b_remelt = SugarStream(
            brix=remelt_brix, purity=b_cent.sugar_purity,
            flow_lb_per_hr=b_remelt_flow,
            temp_deg_F=b_magma_temp_F, pressure_psia=14.7, level_ft=0,
        )

        c_magma_total_flow = c_cent.crystals_to_sugar_lb_hr / (c_magma_brix / 100.0)
        c_to_b_flow        = c_magma_total_flow * (1.0 - c_remelt_pct / 100.0)
        c_remelt_flow      = (c_magma_total_flow * (c_remelt_pct / 100.0)
                              * (c_magma_brix / 100.0) / (remelt_brix / 100.0))
        c_magma_to_b = SugarStream(
            brix=c_magma_brix, purity=c_cent.sugar_purity,
            flow_lb_per_hr=c_to_b_flow,
            temp_deg_F=c_magma_temp_F, pressure_psia=14.7, level_ft=0,
        )
        c_remelt = SugarStream(
            brix=remelt_brix, purity=c_cent.sugar_purity,
            flow_lb_per_hr=c_remelt_flow,
            temp_deg_F=c_magma_temp_F, pressure_psia=14.7, level_ft=0,
        )

        converged = (abs(b_to_a_flow - b_prev)                  < tol and
                     abs(c_to_b_flow - c_prev)                  < tol and
                     abs(seed_stream.flow_lb_per_hr - seed_prev) < tol)
        if converged:
            break

    return {
        'scheme':                    'Three Boiling Double Magma',
        'iterations':                iteration + 1,
        'converged':                 converged,
        # Feed reference (used by summary / station_detail)
        'syrup':                     syrup,
        # A grade
        'a_pan':                     a_pan,
        'a_centrifugal':             a_cent,
        'a_mol_to_b_lb_hr':          a_mol.flow_lb_per_hr * (1.0 - a_mol_to_seed_pct),
        'a_mol_to_seed_lb_hr':       a_mol.flow_lb_per_hr * a_mol_to_seed_pct,
        # B grade
        'b_pan':                     b_pan,
        'b_centrifugal':             b_cent,
        'b_magma_brix':              b_magma_brix,
        'b_magma_purity':            b_cent.sugar_purity,
        'b_magma_temp_F':            b_magma_temp_F,
        'b_magma_total_flow_lb_hr':  b_magma_total_flow,
        'b_magma_to_a_flow_lb_hr':   b_to_a_flow,
        'b_remelt_flow_lb_hr':       b_remelt_flow,
        'b_remelt_brix':             remelt_brix,
        'b_mol_to_c_lb_hr':          b_mol.flow_lb_per_hr * (1.0 - b_mol_to_seed_pct),
        'b_mol_to_seed_lb_hr':       b_mol.flow_lb_per_hr * b_mol_to_seed_pct,
        # Seed pan
        'seed_pan':                  seed_pan,
        'seed_massecuite_flow_lb_hr': seed_pan.massecuite_flow_lb_hr,
        # C grade
        'c_pan':                     c_pan,
        'c_centrifugal':             c_cent,
        'c_magma_brix':              c_magma_brix,
        'c_magma_purity':            c_cent.sugar_purity,
        'c_magma_temp_F':            c_magma_temp_F,
        'c_magma_total_flow_lb_hr':  c_magma_total_flow,
        'c_magma_to_b_flow_lb_hr':   c_to_b_flow,
        'c_remelt_flow_lb_hr':       c_remelt_flow,
        'c_remelt_brix':             remelt_brix,
        # Terminal stream
        'final_molasses':            c_cent.molasses_stream,
    }


# ---------------------------------------------------------------------------
# Summary helpers
# ---------------------------------------------------------------------------

def summary(result):
    """
    Return a flat dict of key flows and compositions for all grades.
    All values are floats except 'scheme' (str) and 'converged' (bool).
    """
    ap, ac = result['a_pan'], result['a_centrifugal']
    bp, bc = result['b_pan'], result['b_centrifugal']
    sp     = result['seed_pan']
    cp, cc = result['c_pan'], result['c_centrifugal']
    fm     = result['final_molasses']

    syr = result['syrup']
    # sucrose crystal true density ~99 lb/ft³; bulk on conveyor ~50-55 lb/ft³
    _SUGAR_DENSITY = 99.0

    return {
        'scheme':                           result['scheme'],
        'iterations':                       result['iterations'],
        'converged':                        result['converged'],
        # ---- Overview / Factory balance ----
        'overview_syrup_flow_lb_hr':        syr.flow_lb_per_hr,
        'overview_syrup_brix':              syr.brix,
        'overview_syrup_purity':            syr.purity,
        'overview_syrup_temp_F':            syr.temp_deg_F,
        'overview_a_sugar_to_warehouse_lb_hr': ac.sugar_wet_lb_hr,
        'overview_a_sugar_pol':             ac.sugar_pol,
        'overview_a_sugar_purity_pct':      ac.sugar_purity,
        'overview_a_sugar_moisture_pct':    ac.sugar_moisture,
        'overview_a_sugar_ft3_hr':          ac.sugar_wet_lb_hr / _SUGAR_DENSITY,
        'overview_b_magma_total_lb_hr':     result['b_magma_total_flow_lb_hr'],
        'overview_b_magma_to_a_pan_lb_hr':  result['b_magma_to_a_flow_lb_hr'],
        'overview_b_remelt_to_syrup_lb_hr': result['b_remelt_flow_lb_hr'],
        'overview_c_magma_total_lb_hr':     result['c_magma_total_flow_lb_hr'],
        'overview_c_magma_to_b_pan_lb_hr':  result['c_magma_to_b_flow_lb_hr'],
        'overview_c_remelt_to_syrup_lb_hr': result['c_remelt_flow_lb_hr'],
        'overview_final_molasses_lb_hr':    fm.flow_lb_per_hr,
        'overview_final_molasses_brix':     fm.brix,
        'overview_final_molasses_purity':   fm.purity,
        'overview_total_steam_lb_hr':       (ap.steam_flow_lb_hr + bp.steam_flow_lb_hr
                                             + sp.steam_flow_lb_hr + cp.steam_flow_lb_hr),
        'overview_total_water_evap_lb_hr':  (ap.water_evaporated_lb_hr + bp.water_evaporated_lb_hr
                                             + sp.water_evaporated_lb_hr + cp.water_evaporated_lb_hr),
        # ---- A grade ----
        'A_feed_flow_lb_hr':                ap.feed_flow_lb_hr,
        'A_feed_solids_lb_hr':              ap.feed_solids_lb_hr,
        'A_massecuite_flow_lb_hr':          ap.massecuite_flow_lb_hr,
        'A_masse_brix':                     ap.masse_brix,
        'A_ml_purity':                      ap.ml_purity,
        'A_masse_purity':                   ap.massecuite.masse_purity,
        'A_crystal_content_pct':            ap.massecuite.crystal_content,
        'A_massecuite_temp_F':              ap.massecuite.massecuite_temp,
        'A_water_evaporated_lb_hr':         ap.water_evaporated_lb_hr,
        'A_steam_type':                     ap.steam_type,
        'A_steam_flow_lb_hr':               ap.steam_flow_lb_hr,
        'A_steam_to_evap_ratio':            ap.steam_to_evaporation_ratio,
        'A_heating_surface_ft2':            ap.heating_surface_ft2,
        'A_delta_T_F':                      ap.delta_T,
        'A_U_btu_hr_ft2_F':                 ap.U_btu_hr_ft2_F,
        'A_sugar_wet_lb_hr':                ac.sugar_wet_lb_hr,
        'A_sugar_purity_pct':               ac.sugar_purity,
        'A_wash_water_lb_hr':               ac.wash_water_lb_hr,
        'A_molasses_flow_lb_hr':            ac.molasses_flow_lb_hr,
        'A_molasses_to_b_pan_lb_hr':        result['a_mol_to_b_lb_hr'],
        'A_molasses_to_seed_lb_hr':         result['a_mol_to_seed_lb_hr'],
        'A_molasses_brix':                  ac.molasses_brix,
        'A_molasses_purity':                ac.molasses_purity,
        'A_purity_rise_pct':                ac.molasses_purity - ap.ml_purity,
        'A_molasses_density_lb_gal':        ac.molasses_density_lb_gal,
        'A_molasses_flow_gal_min':          ac.molasses_flow_gal_min,
        'A_station_crystal_yield_pct_brx':   ac.station_crystal_yield_pct_brix,
        'A_station_crystal_yield_pct_masse':  ac.station_crystal_yield_pct_masse,
        # ---- B grade ----
        'B_magma_total_lb_hr':              result['b_magma_total_flow_lb_hr'],
        'B_magma_to_A_pan_lb_hr':           result['b_magma_to_a_flow_lb_hr'],
        'B_remelt_to_syrup_lb_hr':          result['b_remelt_flow_lb_hr'],
        'B_remelt_brix':                    result['b_remelt_brix'],
        'B_magma_brix':                     result['b_magma_brix'],
        'B_magma_purity':                   result['b_magma_purity'],
        'B_feed_flow_lb_hr':                bp.feed_flow_lb_hr,
        'B_feed_solids_lb_hr':              bp.feed_solids_lb_hr,
        'B_massecuite_flow_lb_hr':          bp.massecuite_flow_lb_hr,
        'B_masse_brix':                     bp.masse_brix,
        'B_ml_purity':                      bp.ml_purity,
        'B_masse_purity':                   bp.massecuite.masse_purity,
        'B_crystal_content_pct':            bp.massecuite.crystal_content,
        'B_massecuite_temp_F':              bp.massecuite.massecuite_temp,
        'B_water_evaporated_lb_hr':         bp.water_evaporated_lb_hr,
        'B_steam_type':                     bp.steam_type,
        'B_steam_flow_lb_hr':               bp.steam_flow_lb_hr,
        'B_steam_to_evap_ratio':            bp.steam_to_evaporation_ratio,
        'B_heating_surface_ft2':            bp.heating_surface_ft2,
        'B_delta_T_F':                      bp.delta_T,
        'B_U_btu_hr_ft2_F':                 bp.U_btu_hr_ft2_F,
        'B_sugar_wet_lb_hr':                bc.sugar_wet_lb_hr,
        'B_sugar_purity_pct':               bc.sugar_purity,
        'B_wash_water_lb_hr':               bc.wash_water_lb_hr,
        'B_molasses_flow_lb_hr':            bc.molasses_flow_lb_hr,
        'B_molasses_to_c_pan_lb_hr':        result['b_mol_to_c_lb_hr'],
        'B_molasses_to_seed_lb_hr':         result['b_mol_to_seed_lb_hr'],
        'B_molasses_brix':                  bc.molasses_brix,
        'B_molasses_purity':                bc.molasses_purity,
        'B_purity_rise_pct':                bc.molasses_purity - bp.ml_purity,
        'B_molasses_density_lb_gal':        bc.molasses_density_lb_gal,
        'B_molasses_flow_gal_min':          bc.molasses_flow_gal_min,
        'B_station_crystal_yield_pct_brx':   bc.station_crystal_yield_pct_brix,
        'B_station_crystal_yield_pct_masse':  bc.station_crystal_yield_pct_masse,
        # ---- Seed pan ----
        'Seed_feed_flow_lb_hr':             sp.feed_flow_lb_hr,
        'Seed_feed_solids_lb_hr':           sp.feed_solids_lb_hr,
        'Seed_massecuite_flow_lb_hr':       sp.massecuite_flow_lb_hr,
        'Seed_masse_brix':                  sp.masse_brix,
        'Seed_ml_purity':                   sp.ml_purity,
        'Seed_masse_purity':                sp.massecuite.masse_purity,
        'Seed_crystal_content_pct':         sp.massecuite.crystal_content,
        'Seed_massecuite_temp_F':           sp.massecuite.massecuite_temp,
        'Seed_water_evaporated_lb_hr':      sp.water_evaporated_lb_hr,
        'Seed_steam_type':                  sp.steam_type,
        'Seed_steam_flow_lb_hr':            sp.steam_flow_lb_hr,
        'Seed_steam_to_evap_ratio':         sp.steam_to_evaporation_ratio,
        'Seed_heating_surface_ft2':         sp.heating_surface_ft2,
        'Seed_delta_T_F':                   sp.delta_T,
        'Seed_U_btu_hr_ft2_F':              sp.U_btu_hr_ft2_F,
        # ---- C grade ----
        'C_magma_total_lb_hr':              result['c_magma_total_flow_lb_hr'],
        'C_magma_to_B_pan_lb_hr':           result['c_magma_to_b_flow_lb_hr'],
        'C_remelt_to_syrup_lb_hr':          result['c_remelt_flow_lb_hr'],
        'C_remelt_brix':                    result['c_remelt_brix'],
        'C_magma_brix':                     result['c_magma_brix'],
        'C_magma_purity':                   result['c_magma_purity'],
        'C_feed_flow_lb_hr':                cp.feed_flow_lb_hr,
        'C_feed_solids_lb_hr':              cp.feed_solids_lb_hr,
        'C_massecuite_flow_lb_hr':          cp.massecuite_flow_lb_hr,
        'C_masse_brix':                     cp.masse_brix,
        'C_ml_purity':                      cp.ml_purity,
        'C_masse_purity':                   cp.massecuite.masse_purity,
        'C_crystal_content_pct':            cp.massecuite.crystal_content,
        'C_massecuite_temp_F':              cp.massecuite.massecuite_temp,
        'C_water_evaporated_lb_hr':         cp.water_evaporated_lb_hr,
        'C_steam_type':                     cp.steam_type,
        'C_steam_flow_lb_hr':               cp.steam_flow_lb_hr,
        'C_steam_to_evap_ratio':            cp.steam_to_evaporation_ratio,
        'C_heating_surface_ft2':            cp.heating_surface_ft2,
        'C_delta_T_F':                      cp.delta_T,
        'C_U_btu_hr_ft2_F':                 cp.U_btu_hr_ft2_F,
        'C_sugar_wet_lb_hr':                cc.sugar_wet_lb_hr,
        'C_sugar_purity_pct':               cc.sugar_purity,
        'C_wash_water_lb_hr':               cc.wash_water_lb_hr,
        'C_molasses_flow_lb_hr':            cc.molasses_flow_lb_hr,
        'C_molasses_brix':                  cc.molasses_brix,
        'C_molasses_purity':                cc.molasses_purity,
        'C_purity_rise_pct':                cc.molasses_purity - cp.ml_purity,
        'C_molasses_density_lb_gal':        cc.molasses_density_lb_gal,
        'C_molasses_flow_gal_min':          cc.molasses_flow_gal_min,
        'C_station_crystal_yield_pct_brx':   cc.station_crystal_yield_pct_brix,
        'C_station_crystal_yield_pct_masse':  cc.station_crystal_yield_pct_masse,
        # ---- Final molasses ----
        'final_molasses_flow_lb_hr':        fm.flow_lb_per_hr,
        'final_molasses_brix':              fm.brix,
        'final_molasses_purity':            fm.purity,
        # ---- Totals ----
        'total_sugar_wet_lb_hr':  ac.sugar_wet_lb_hr + bc.sugar_wet_lb_hr + cc.sugar_wet_lb_hr,
        'total_steam_lb_hr':      (ap.steam_flow_lb_hr + bp.steam_flow_lb_hr
                                   + sp.steam_flow_lb_hr + cp.steam_flow_lb_hr),
        'total_water_evap_lb_hr': (ap.water_evaporated_lb_hr + bp.water_evaporated_lb_hr
                                   + sp.water_evaporated_lb_hr + cp.water_evaporated_lb_hr),
    }


def display_summary(result):
    """Print a formatted table of the scheme summary."""
    s = summary(result)

    print(f"\n{'='*62}")
    print(f"  {s['scheme']}")
    print(f"  Converged: {s['converged']}   Iterations: {s['iterations']}")
    print(f"{'='*62}")

    sections = [
        ('OVERVIEW',       'overview_'),
        ('A GRADE',        'A_'),
        ('B GRADE',        'B_'),
        ('SEED PAN',       'Seed_'),
        ('C GRADE',        'C_'),
        ('FINAL MOLASSES', 'final_'),
        ('TOTALS',         'total_'),
    ]
    skip = {'scheme', 'iterations', 'converged'}

    for title, prefix in sections:
        keys = [k for k in s if k.startswith(prefix) and k not in skip]
        if not keys:
            continue
        print(f"\n  --- {title} ---")
        for k in keys:
            label = k[len(prefix):].replace('_', ' ')
            v = s[k]
            if isinstance(v, str):
                print(f"    {label:<40}: {v}")
            else:
                print(f"    {label:<40}: {v:>12,.3f}")


# ---------------------------------------------------------------------------
# Station detail (entering / leaving per unit)
# ---------------------------------------------------------------------------

def station_detail(result):
    """
    Return a list of station dicts with entering and leaving stream tables.
    Each entry: {'name': str, 'entering': [...], 'leaving': [...]}
    Stream row keys: stream, flow_lb_hr, brix, purity, temp_F,
                     water_lb_hr, pol_lb_hr, solids_lb_hr, balance
    balance=False marks informational rows excluded from subtotals
    (magma recycle shown for context after wet sugar row, steam excluded as energy stream).
    """
    ap, ac = result['a_pan'], result['a_centrifugal']
    bp, bc = result['b_pan'], result['b_centrifugal']
    sp     = result['seed_pan']
    cp, cc = result['c_pan'], result['c_centrifugal']
    syr    = result['syrup']

    def row(name, flow, brix=None, purity=None, temp=None, balance=True):
        """Build a stream row with auto-computed water/pol/solids and pol_pct.
            brix (%)   — solids fraction: sugar rows use 100 − moisture
            purity (%) — AP = pol/solids; sugar rows use the specified sugar_purity
            pol_pct    = brix × purity / 100
            pol_lb_hr  = flow × pol_pct / 100 = crystals × purity / 100
        """
        nl = name.lower()
        if 'steam' in nl:
            w = p = s = pol_pct = None
            balance = False
        elif flow is not None and brix is not None and purity is not None:
            s       = flow * brix / 100.0
            w       = flow - s
            pol_pct = brix * purity / 100.0
            p       = flow * pol_pct / 100.0
        elif flow is not None and brix is not None:
            s       = flow * brix / 100.0
            w       = flow - s
            p       = None
            pol_pct = None
        elif flow is not None:
            if any(t in nl for t in ('water', 'evapor', 'wash', 'condensate')):
                w, p, s, pol_pct = flow, 0.0, 0.0, 0.0
            else:
                w = p = s = pol_pct = None
        else:
            w = p = s = pol_pct = None
        return {
            'stream': name, 'flow_lb_hr': flow,
            'brix': brix, 'purity': purity, 'temp_F': temp,
            'water_lb_hr': w, 'pol_lb_hr': p, 'solids_lb_hr': s,
            'pol_pct': pol_pct,
            'balance': balance,
        }

    # Derived stream props reused across stations
    a_mol_brix = ac.molasses_brix;  a_mol_pur = ac.molasses_purity
    a_mol_temp = ap.massecuite.massecuite_temp
    b_mol_brix = bc.molasses_brix;  b_mol_pur = bc.molasses_purity
    b_mol_temp = bp.massecuite.massecuite_temp
    seed_brix  = sp.masse_brix;     seed_pur  = sp.massecuite.masse_purity
    seed_temp  = sp.massecuite.massecuite_temp
    b_mg_brix  = result['b_magma_brix']; b_mg_temp = result['b_magma_temp_F']
    c_mg_brix  = result['c_magma_brix']; c_mg_temp = result['c_magma_temp_F']

    return [
        {'name': 'A PAN',
         'entering': [
             row('Syrup',              syr.flow_lb_per_hr,                    syr.brix,    syr.purity,    syr.temp_deg_F),
             row('B Magma (footing)', result['b_magma_to_a_flow_lb_hr'],     b_mg_brix,  bc.sugar_purity, b_mg_temp),
             row('B Remelt (syrup)',  result['b_remelt_flow_lb_hr'],          result['b_remelt_brix'], bc.sugar_purity, b_mg_temp),
             row('C Remelt (syrup)',  result['c_remelt_flow_lb_hr'],          result['c_remelt_brix'], cc.sugar_purity, c_mg_temp),
         ],
         'leaving': [
             row('A Massecuite',     ap.massecuite_flow_lb_hr, ap.masse_brix, ap.massecuite.masse_purity, ap.massecuite.massecuite_temp),
             row('Water Evaporated', ap.water_evaporated_lb_hr),
             row('Steam Consumed',   ap.steam_flow_lb_hr),
         ]},

        {'name': 'A CENTRIFUGAL',
         'entering': [
             row('A Massecuite',                     ap.massecuite_flow_lb_hr, ap.masse_brix, ap.massecuite.masse_purity, ap.massecuite.massecuite_temp),
             row('  ↳ Water Already in Massecuite',  ap.massecuite_flow_lb_hr * (1 - ap.masse_brix/100), balance=False),
             row('Wash Water Added',                 ac.wash_water_lb_hr),
         ],
         'leaving': [
             row('A Sugar → Warehouse', ac.sugar_wet_lb_hr,
                 ac.sugar_brix, ac.sugar_purity),
             row('A Molasses → B Pan',    result['a_mol_to_b_lb_hr'],    a_mol_brix, a_mol_pur, a_mol_temp),
             row('A Molasses → Seed Pan', result['a_mol_to_seed_lb_hr'], a_mol_brix, a_mol_pur, a_mol_temp),
         ]},

        {'name': 'B PAN',
         'entering': [
             row('A Molasses',       result['a_mol_to_b_lb_hr'],   a_mol_brix, a_mol_pur, a_mol_temp),
             row('C Magma (footing)', result['c_magma_to_b_flow_lb_hr'], c_mg_brix, cc.sugar_purity, c_mg_temp),
         ],
         'leaving': [
             row('B Massecuite',     bp.massecuite_flow_lb_hr, bp.masse_brix, bp.massecuite.masse_purity, bp.massecuite.massecuite_temp),
             row('Water Evaporated', bp.water_evaporated_lb_hr),
             row('Steam Consumed',   bp.steam_flow_lb_hr),
         ]},

        {'name': 'B CENTRIFUGAL',
         'entering': [
             row('B Massecuite',                     bp.massecuite_flow_lb_hr, bp.masse_brix, bp.massecuite.masse_purity, bp.massecuite.massecuite_temp),
             row('  ↳ Water Already in Massecuite',  bp.massecuite_flow_lb_hr * (1 - bp.masse_brix/100), balance=False),
             row('Wash Water Added',                 bc.wash_water_lb_hr),
         ],
         'leaving': [
             row('B Wet Sugar (→ B Magma)', bc.sugar_wet_lb_hr,
                 bc.sugar_brix, bc.sugar_purity),
             row('  + Water Added for Magma (footing)', result['b_magma_total_flow_lb_hr'] - bc.sugar_wet_lb_hr, balance=False),
             row('  + Water Added for Remelt',         result['b_remelt_flow_lb_hr'] - (result['b_magma_total_flow_lb_hr'] - result['b_magma_to_a_flow_lb_hr']), balance=False),
             row('B Molasses → C Pan',    result['b_mol_to_c_lb_hr'],    b_mol_brix, b_mol_pur, b_mol_temp),
             row('B Molasses → Seed Pan', result['b_mol_to_seed_lb_hr'], b_mol_brix, b_mol_pur, b_mol_temp),
             row('  → B Magma to A Pan',  result['b_magma_to_a_flow_lb_hr'], b_mg_brix, result['b_magma_purity'], b_mg_temp, balance=False),
             row('  → B Remelt to Syrup', result['b_remelt_flow_lb_hr'],     result['b_remelt_brix'], result['b_magma_purity'], b_mg_temp, balance=False),
         ]},

        {'name': 'SEED PAN',
         'entering': [
             row('A Molasses', result['a_mol_to_seed_lb_hr'], a_mol_brix, a_mol_pur, a_mol_temp),
             row('B Molasses', result['b_mol_to_seed_lb_hr'], b_mol_brix, b_mol_pur, b_mol_temp),
         ],
         'leaving': [
             row('Seed Massecuite → C Pan', sp.massecuite_flow_lb_hr, seed_brix, seed_pur, seed_temp),
             row('Water Evaporated',        sp.water_evaporated_lb_hr),
             row('Steam Consumed',          sp.steam_flow_lb_hr),
         ]},

        {'name': 'C PAN',
         'entering': [
             row('B Molasses',      result['b_mol_to_c_lb_hr'], b_mol_brix, b_mol_pur, b_mol_temp),
             row('Seed Massecuite', sp.massecuite_flow_lb_hr,   seed_brix,  seed_pur,  seed_temp),
         ],
         'leaving': [
             row('C Massecuite',    cp.massecuite_flow_lb_hr, cp.masse_brix, cp.massecuite.masse_purity, cp.massecuite.massecuite_temp),
             row('Water Evaporated', cp.water_evaporated_lb_hr),
             row('Steam Consumed',   cp.steam_flow_lb_hr),
         ]},

        {'name': 'C CENTRIFUGAL',
         'entering': [
             row('C Massecuite',                     cp.massecuite_flow_lb_hr, cp.masse_brix, cp.massecuite.masse_purity, cp.massecuite.massecuite_temp),
             row('  ↳ Water Already in Massecuite',  cp.massecuite_flow_lb_hr * (1 - cp.masse_brix/100), balance=False),
             row('Wash Water Added',                 cc.wash_water_lb_hr),
         ],
         'leaving': [
             row('C Wet Sugar (→ C Magma)', cc.sugar_wet_lb_hr,
                 cc.sugar_brix, cc.sugar_purity),
             row('  + Water Added for Magma (footing)', result['c_magma_total_flow_lb_hr'] - cc.sugar_wet_lb_hr, balance=False),
             row('  + Water Added for Remelt',         result['c_remelt_flow_lb_hr'] - (result['c_magma_total_flow_lb_hr'] - result['c_magma_to_b_flow_lb_hr']), balance=False),
             row('Final Molasses', cc.molasses_flow_lb_hr, cc.molasses_brix, cc.molasses_purity, cp.massecuite.massecuite_temp),
             row('  → C Magma to B Pan',  result['c_magma_to_b_flow_lb_hr'],  c_mg_brix, result['c_magma_purity'], c_mg_temp, balance=False),
             row('  → C Remelt to Syrup', result['c_remelt_flow_lb_hr'],       result['c_remelt_brix'], result['c_magma_purity'], c_mg_temp, balance=False),
         ]},
    ]


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def _infer_units(key):
    if 'lb_hr' in key:        return 'lb/hr'
    if 'btu_hr_ft2' in key:   return 'BTU/hr·ft²·°F'
    if 'ft3' in key:          return 'ft³/hr'
    if 'ft2' in key:          return 'ft²'
    if 'lb_gal' in key:       return 'lb/gal'
    if 'gal_min' in key:      return 'gal/min'
    if 'ratio' in key:        return 'lb/lb'
    if 'temp_F' in key:       return '°F'
    if key.endswith('_F'):    return '°F'
    if 'delta_T' in key:      return '°F'
    if 'pol' in key:          return 'Pol'
    if 'brix' in key:         return 'Brix'
    if 'purity' in key:       return '%'
    if 'pct' in key:          return '%'
    if 'moisture' in key:     return '%'
    if 'steam_type' in key:   return ''
    return ''


def to_excel(result, filepath='boiling_scheme.xlsx'):
    """
    Export the scheme summary to a formatted Excel workbook.
    Sections are color-coded; flows use comma formatting; ratios and
    temperatures use appropriate decimal places.
    Requires openpyxl (already in the project venv).
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    s = summary(result)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Boiling Scheme'

    # ---- Color palette ----
    TITLE_FILL   = PatternFill('solid', fgColor='1F4E79')  # dark navy
    COLHDR_FILL  = PatternFill('solid', fgColor='404040')  # charcoal
    SEC_FILLS = {
        'OVERVIEW':       PatternFill('solid', fgColor='1F4E79'),  # dark navy
        'A GRADE':        PatternFill('solid', fgColor='2E75B6'),  # blue
        'B GRADE':        PatternFill('solid', fgColor='70AD47'),  # green
        'SEED PAN':       PatternFill('solid', fgColor='ED7D31'),  # orange
        'C GRADE':        PatternFill('solid', fgColor='C00000'),  # red
        'FINAL MOLASSES': PatternFill('solid', fgColor='7030A0'),  # purple
        'TOTALS':         PatternFill('solid', fgColor='404040'),  # charcoal
    }
    ALT_FILLS = {
        'OVERVIEW':       PatternFill('solid', fgColor='D6DCE4'),
        'A GRADE':        PatternFill('solid', fgColor='D6E4F0'),
        'B GRADE':        PatternFill('solid', fgColor='E2EFDA'),
        'SEED PAN':       PatternFill('solid', fgColor='FCE4D6'),
        'C GRADE':        PatternFill('solid', fgColor='FFDBD8'),
        'FINAL MOLASSES': PatternFill('solid', fgColor='EFE0FF'),
        'TOTALS':         PatternFill('solid', fgColor='EDEDED'),
    }
    WHITE = PatternFill('solid', fgColor='FFFFFF')

    thin   = Side(style='thin', color='BFBFBF')
    bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')
    left   = Alignment(horizontal='left',   vertical='center')
    right  = Alignment(horizontal='right',  vertical='center')

    def hdr_font(sz=11): return Font(bold=True, color='FFFFFF', size=sz)
    def body_font():     return Font(size=10)

    # ---- Title rows ----
    ws.merge_cells('A1:C1')
    c = ws['A1']
    c.value, c.font, c.fill, c.alignment = (s['scheme'], hdr_font(13),
                                              TITLE_FILL, center)

    ws.merge_cells('A2:C2')
    c = ws['A2']
    c.value = f"Converged: {s['converged']}    Iterations: {s['iterations']}"
    c.font, c.fill, c.alignment = hdr_font(10), TITLE_FILL, center

    # ---- Column headers ----
    row = 3
    for col, text in enumerate(['Parameter', 'Value', 'Units'], start=1):
        c = ws.cell(row=row, column=col, value=text)
        c.font, c.fill, c.alignment, c.border = (hdr_font(10), COLHDR_FILL,
                                                   center, bdr)
    row += 1

    # ---- Data sections ----
    sections = [
        ('OVERVIEW',       'overview_'),
        ('A GRADE',        'A_'),
        ('B GRADE',        'B_'),
        ('SEED PAN',       'Seed_'),
        ('C GRADE',        'C_'),
        ('FINAL MOLASSES', 'final_'),
        ('TOTALS',         'total_'),
    ]
    skip = {'scheme', 'iterations', 'converged'}

    for title, prefix in sections:
        keys = [k for k in s if k.startswith(prefix) and k not in skip]
        if not keys:
            continue

        sec_fill = SEC_FILLS[title]
        alt_fill = ALT_FILLS[title]

        # Section header spanning all three columns
        ws.merge_cells(f'A{row}:C{row}')
        ws.row_dimensions[row].height = 16
        c = ws.cell(row=row, column=1, value=title)
        c.font, c.fill, c.alignment = hdr_font(11), sec_fill, left
        for col in (2, 3):
            ws.cell(row=row, column=col).fill = sec_fill
        row += 1

        alt = False
        for k in keys:
            label = k[len(prefix):].replace('_', ' ').title()
            v     = s[k]
            units = _infer_units(k)
            fill  = alt_fill if alt else WHITE

            lc = ws.cell(row=row, column=1, value=label)
            vc = ws.cell(row=row, column=2, value=v)
            uc = ws.cell(row=row, column=3, value=units)

            for cell, align in ((lc, left), (vc, right), (uc, left)):
                cell.font, cell.fill, cell.alignment, cell.border = (
                    body_font(), fill, align, bdr)

            if isinstance(v, float):
                if units == 'lb/hr':     vc.number_format = '#,##0'
                elif units == '%':       vc.number_format = '0.00'
                elif units == 'Brix':    vc.number_format = '0.0'
                elif units == '°F':      vc.number_format = '0.0'
                elif units == 'lb/lb':   vc.number_format = '0.000'
                elif units == 'lb/gal':  vc.number_format = '0.000'
                elif units == 'gal/min': vc.number_format = '#,##0.0'
                elif 'btu' in k:         vc.number_format = '0.0'
                else:                    vc.number_format = '#,##0.00'

            alt = not alt
            row += 1

        row += 1  # blank separator between sections

    # ---- Column widths and freeze ----
    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.freeze_panes = 'A4'

    # ---- Station Detail sheet ----
    ws2 = wb.create_sheet('Station Detail')

    STATION_COLORS = {
        'A PAN':         ('1F4E79', 'D6E4F0'),
        'A CENTRIFUGAL': ('2E75B6', 'DEEAF1'),
        'B PAN':         ('375623', 'E2EFDA'),
        'B CENTRIFUGAL': ('70AD47', 'EDF7E3'),
        'SEED PAN':      ('843C0C', 'FCE4D6'),
        'C PAN':         ('843150', 'FFDBD8'),
        'C CENTRIFUGAL': ('C00000', 'FFE7E7'),
    }
    IN_COLOR  = ('375623', 'EDF7E3')   # dark/light green for entering
    OUT_COLOR = ('843150', 'FFE7E7')   # dark/light red for leaving

    col_headers = ['Stream', 'Flow (lb/hr)', 'Brix', 'Purity (%)', 'Pol (%)', 'Temp (°F)',
                   'Water (lb/hr)', 'Pol (lb/hr)', 'Solids (lb/hr)']
    col_widths2  = [34, 16, 10, 12, 10, 12, 16, 16, 16]
    col_fmts     = [None, '#,##0', '0.0', '0.00', '0.00', '0.0', '#,##0', '#,##0', '#,##0']
    NCOLS2 = len(col_headers)
    last_col_letter = get_column_letter(NCOLS2)

    # Write column headers row
    r2 = 1
    for ci, (hdr, w) in enumerate(zip(col_headers, col_widths2), start=1):
        c = ws2.cell(row=r2, column=ci, value=hdr)
        c.font = Font(bold=True, color='FFFFFF', size=10)
        c.fill = PatternFill('solid', fgColor='404040')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = bdr
        ws2.column_dimensions[get_column_letter(ci)].width = w
    r2 += 1

    SUBTOTAL_FILL_IN  = PatternFill('solid', fgColor='C6EFCE')   # soft green
    SUBTOTAL_FILL_OUT = PatternFill('solid', fgColor='FFCCCC')    # soft red
    SUBTOTAL_FONT     = Font(bold=True, size=10)
    INFO_FONT         = Font(italic=True, color='808080', size=10) # grey italic for balance=False rows

    for station in station_detail(result):
        name = station['name']
        sta_dark, _ = STATION_COLORS.get(name, ('404040', 'F2F2F2'))

        # Station header spanning all columns
        ws2.merge_cells(f'A{r2}:{last_col_letter}{r2}')
        c = ws2.cell(row=r2, column=1, value=name)
        c.font = Font(bold=True, color='FFFFFF', size=11)
        c.fill = PatternFill('solid', fgColor=sta_dark)
        c.alignment = Alignment(horizontal='left', vertical='center')
        c.border = bdr
        for ci in range(2, NCOLS2 + 1):
            ws2.cell(row=r2, column=ci).fill = PatternFill('solid', fgColor=sta_dark)
            ws2.cell(row=r2, column=ci).border = bdr
        ws2.row_dimensions[r2].height = 15
        r2 += 1

        for direction, streams in [('ENTERING', station['entering']),
                                    ('LEAVING',  station['leaving'])]:
            dir_dark, dir_light = IN_COLOR if direction == 'ENTERING' else OUT_COLOR
            sub_fill = SUBTOTAL_FILL_IN if direction == 'ENTERING' else SUBTOTAL_FILL_OUT

            # Direction sub-header spanning all columns
            ws2.merge_cells(f'A{r2}:{last_col_letter}{r2}')
            c = ws2.cell(row=r2, column=1, value=f'  {direction}')
            c.font = Font(bold=True, color='FFFFFF', size=10)
            c.fill = PatternFill('solid', fgColor=dir_dark)
            c.alignment = Alignment(horizontal='left', vertical='center')
            for ci in range(2, NCOLS2 + 1):
                ws2.cell(row=r2, column=ci).fill = PatternFill('solid', fgColor=dir_dark)
            r2 += 1

            alt = False
            for sr in streams:
                is_info = not sr.get('balance', True)
                fill = PatternFill('solid', fgColor=dir_light if alt else 'FFFFFF')
                vals = [sr['stream'], sr['flow_lb_hr'],
                        sr['brix'], sr['purity'], sr.get('pol_pct'), sr['temp_F'],
                        sr.get('water_lb_hr'), sr.get('pol_lb_hr'), sr.get('solids_lb_hr')]
                for ci, (v, fmt) in enumerate(zip(vals, col_fmts), start=1):
                    c = ws2.cell(row=r2, column=ci, value=v)
                    c.font  = INFO_FONT if is_info else Font(size=10)
                    c.fill  = fill
                    c.border = bdr
                    c.alignment = Alignment(
                        horizontal='right' if ci > 1 else 'left',
                        vertical='center')
                    if v is not None and fmt and isinstance(v, (int, float)):
                        c.number_format = fmt
                    if v is None:
                        c.value = '—'
                        c.alignment = Alignment(horizontal='center', vertical='center')
                if not is_info:
                    alt = not alt
                r2 += 1

            # Subtotal row (sum only balance=True rows)
            bal = [sr for sr in streams if sr.get('balance', True)
                   and sr.get('water_lb_hr') is not None]  # skip steam rows
            if bal:
                sub_vals = {
                    'flow':   sum(sr['flow_lb_hr']   for sr in bal if sr['flow_lb_hr']   is not None),
                    'water':  sum(sr['water_lb_hr']  for sr in bal if sr['water_lb_hr']  is not None),
                    'pol':    sum(sr['pol_lb_hr']     for sr in bal if sr['pol_lb_hr']    is not None),
                    'solids': sum(sr['solids_lb_hr'] for sr in bal if sr['solids_lb_hr'] is not None),
                }
                sub_row = [f'  SUBTOTAL {direction}', sub_vals['flow'],
                           None, None, None, None,          # Brix, Purity, Pol%, Temp
                           sub_vals['water'], sub_vals['pol'], sub_vals['solids']]
                for ci, (v, fmt) in enumerate(zip(sub_row, col_fmts), start=1):
                    c = ws2.cell(row=r2, column=ci, value=v)
                    c.font  = SUBTOTAL_FONT
                    c.fill  = sub_fill
                    c.border = bdr
                    c.alignment = Alignment(
                        horizontal='right' if ci > 1 else 'left',
                        vertical='center')
                    if v is not None and fmt and isinstance(v, (int, float)):
                        c.number_format = fmt
                    if v is None:
                        c.value = '—'
                        c.alignment = Alignment(horizontal='center', vertical='center')
                r2 += 1

        r2 += 1  # blank row between stations

    ws2.freeze_panes = 'A2'

    # ---- System Balance sheet ----
    ws3 = wb.create_sheet('System Balance')

    syr_ = result['syrup']
    ac_  = result['a_centrifugal']
    cc_  = result['c_centrifugal']

    pol_in           = syr_.flow_lb_per_hr * syr_.brix/100 * syr_.purity/100
    pol_out_sugar    = ac_.sugar_pol_lb_hr
    pol_out_molasses = cc_.pol_to_molasses_lb_hr
    pol_out_total    = pol_out_sugar + pol_out_molasses
    pol_imbalance    = pol_in - pol_out_total

    solids_in           = syr_.flow_lb_per_hr * syr_.brix/100
    solids_out_sugar    = ac_.crystals_to_sugar_lb_hr
    solids_out_molasses = cc_.molasses_solids_lb_hr
    solids_out_total    = solids_out_sugar + solids_out_molasses
    solids_imbalance    = solids_in - solids_out_total

    bal_headers  = ['Component', 'IN (lb/hr)', 'OUT — A Sugar', 'OUT — Final Molasses', 'Total OUT', 'Imbalance (lb/hr)', 'Imbalance (%)']
    bal_widths   = [22, 18, 18, 22, 18, 20, 16]
    bal_data = [
        ('Pol',    pol_in,    pol_out_sugar,    pol_out_molasses,    pol_out_total,    pol_imbalance,    abs(pol_imbalance)/pol_in*100 if pol_in else 0),
        ('Solids', solids_in, solids_out_sugar, solids_out_molasses, solids_out_total, solids_imbalance, abs(solids_imbalance)/solids_in*100 if solids_in else 0),
    ]

    # Title
    ws3.merge_cells(f'A1:{get_column_letter(len(bal_headers))}1')
    c = ws3.cell(row=1, column=1, value='System Balance Check — Pol and Solids')
    c.font      = Font(bold=True, color='FFFFFF', size=12)
    c.fill      = PatternFill('solid', fgColor='1F4E79')
    c.alignment = Alignment(horizontal='center', vertical='center')
    for ci in range(2, len(bal_headers)+1):
        ws3.cell(row=1, column=ci).fill = PatternFill('solid', fgColor='1F4E79')

    ws3.merge_cells(f'A2:{get_column_letter(len(bal_headers))}2')
    c = ws3.cell(row=1, column=1, value='System Balance Check — Pol and Solids')
    c = ws3.cell(row=2, column=1, value='Only A Sugar (warehouse) and Final Molasses leave the system.  All magma / seed recycles are internal.')
    c.font      = Font(italic=True, size=10, color='404040')
    c.alignment = Alignment(horizontal='left', vertical='center')

    # Column headers
    for ci, (hdr, w) in enumerate(zip(bal_headers, bal_widths), start=1):
        c = ws3.cell(row=3, column=ci, value=hdr)
        c.font      = Font(bold=True, color='FFFFFF', size=10)
        c.fill      = PatternFill('solid', fgColor='404040')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border    = bdr
        ws3.column_dimensions[get_column_letter(ci)].width = w

    # Data rows
    BAL_FILLS = [
        PatternFill('solid', fgColor='D6E4F0'),
        PatternFill('solid', fgColor='E2EFDA'),
    ]
    IMBAL_GOOD = PatternFill('solid', fgColor='C6EFCE')   # green — small imbalance
    IMBAL_WARN = PatternFill('solid', fgColor='FFEB9C')   # yellow — warn > 0.1%
    IMBAL_BAD  = PatternFill('solid', fgColor='FFC7CE')   # red — > 1%

    for ri, (row_data, row_fill) in enumerate(zip(bal_data, BAL_FILLS), start=4):
        comp, val_in, val_sugar, val_mol, val_total, val_imbal, val_pct = row_data
        row_vals = [comp, val_in, val_sugar, val_mol, val_total, val_imbal, val_pct]
        row_fmts = [None, '#,##0.0', '#,##0.0', '#,##0.0', '#,##0.0', '+#,##0.00;-#,##0.00;0.00', '0.0000%']

        for ci, (v, fmt) in enumerate(zip(row_vals, row_fmts), start=1):
            c = ws3.cell(row=ri, column=ci, value=v)
            c.font      = Font(bold=(ci == 1), size=10)
            c.fill      = row_fill
            c.border    = bdr
            c.alignment = Alignment(horizontal='left' if ci == 1 else 'right', vertical='center')
            if fmt and isinstance(v, (int, float)):
                c.number_format = fmt

        # Color the imbalance cell based on magnitude
        abs_pct = abs(val_pct)
        imbal_fill = IMBAL_GOOD if abs_pct < 0.01 else (IMBAL_WARN if abs_pct < 1.0 else IMBAL_BAD)
        for ci in (6, 7):
            ws3.cell(row=ri, column=ci).fill = imbal_fill

        ws3.row_dimensions[ri].height = 18

    ws3.freeze_panes = 'A4'

    wb.save(filepath)
    print(f"Saved: {filepath}")
    return filepath


# ---------------------------------------------------------------------------
# Test block
# ---------------------------------------------------------------------------

if __name__ == "__main__":

    syrup = SugarStream(
        brix=65.0, purity=87.0,
        flow_lb_per_hr=300_000,
        temp_deg_F=140.0, pressure_psia=14.7, level_ft=0,
    )

    a_pan_specs = dict(
        heating_surface_ft2=22_000,
        inches_vacuum=23.5, supersaturation=1.20, head_ft=2.0,
        masse_brix=92, ml_purity=70,
        steam_type='V1', heat_loss_factor=0.05,
    )
    b_pan_specs = dict(
        heating_surface_ft2=14_000,
        inches_vacuum=25.0, supersaturation=1.20, head_ft=2.0,
        masse_brix=93, ml_purity=52,
        steam_type='V1', heat_loss_factor=0.15,
    )
    seed_pan_specs = dict(
        heating_surface_ft2=4_000,
        inches_vacuum=26.0, supersaturation=1.20, head_ft=2.0,
        masse_brix=88, ml_purity=40,   # ml_purity must be < expected feed purity (~63%)
        steam_type='V1', heat_loss_factor=0.05,
    )
    c_pan_specs = dict(
        heating_surface_ft2=4_000,
        inches_vacuum=26.5, supersaturation=1.00, head_ft=2.0,
        masse_brix=96, ml_purity=30,
        steam_type='V1', heat_loss_factor=0.05,
    )

    a_cent_specs = dict(
        target_molasses_brix=78.0, purity_rise=1.5,
        sugar_purity=99.7, sugar_moisture=0.2,
    )
    b_cent_specs = dict(
        target_molasses_brix=80.0, purity_rise=8.0,
        sugar_purity=92.0, sugar_moisture=4.0,
    )
    c_cent_specs = dict(
        target_molasses_brix=80.0, purity_rise=3.5,
        sugar_purity=85.0, sugar_moisture=5.0,
    )

    result = three_boiling_double_magma(
        syrup=syrup,
        a_pan_specs=a_pan_specs,
        b_pan_specs=b_pan_specs,
        c_pan_specs=c_pan_specs,
        seed_pan_specs=seed_pan_specs,
        a_cent_specs=a_cent_specs,
        b_cent_specs=b_cent_specs,
        c_cent_specs=c_cent_specs,
        b_magma_brix=92.0,
        c_magma_brix=92.0,
        b_magma_temp_F=120.0,
        c_magma_temp_F=120.0,
        a_mol_to_seed_pct=0.10,
        b_mol_to_seed_pct=0.10,
        b_remelt_pct=25.0,
        c_remelt_pct=25.0,
        remelt_brix=65.0,
    )

    display_summary(result)

    print("\n--- Convergence ---")
    print(f"  Iterations      : {result['iterations']}")
    print(f"  B magma to A pan: {result['b_magma_to_a_flow_lb_hr']:,.0f} lb/hr  at {result['b_magma_brix']:.0f} Brix  (remelt: {result['b_remelt_flow_lb_hr']:,.0f} lb/hr at {result['b_remelt_brix']:.0f} Brix)")
    print(f"  C magma to B pan: {result['c_magma_to_b_flow_lb_hr']:,.0f} lb/hr  at {result['c_magma_brix']:.0f} Brix  (remelt: {result['c_remelt_flow_lb_hr']:,.0f} lb/hr at {result['c_remelt_brix']:.0f} Brix)")
    print(f"  Seed massecuite : {result['seed_massecuite_flow_lb_hr']:,.0f} lb/hr  to C pan")

    # ---- System-level pol and solids balance check ----
    # Only A sugar (to warehouse) and final molasses leave the system.
    # All magma and seed loops are internal — they must cancel at steady state.
    _syr = result['syrup']
    _ac  = result['a_centrifugal']
    _cc  = result['c_centrifugal']

    pol_in          = _syr.flow_lb_per_hr * _syr.brix / 100 * _syr.purity / 100
    pol_out_sugar   = _ac.sugar_pol_lb_hr
    pol_out_molasses= _cc.pol_to_molasses_lb_hr
    pol_imbalance   = pol_in - pol_out_sugar - pol_out_molasses

    solids_in          = _syr.flow_lb_per_hr * _syr.brix / 100
    solids_out_sugar   = _ac.crystals_to_sugar_lb_hr   # dry solids in sugar product
    solids_out_molasses= _cc.molasses_solids_lb_hr
    solids_imbalance   = solids_in - solids_out_sugar - solids_out_molasses

    print("\n=== SYSTEM BALANCE CHECK ===")
    print(f"  {'':30}  {'IN':>14}  {'OUT':>14}  {'IMBALANCE':>14}")
    print(f"  {'Pol (lb/hr)':30}  {pol_in:>14,.1f}  {pol_out_sugar+pol_out_molasses:>14,.1f}  {pol_imbalance:>+14.4f}")
    print(f"    A sugar pol               : {pol_out_sugar:>14,.1f}")
    print(f"    Final molasses pol        : {pol_out_molasses:>14,.1f}")
    print(f"  {'Solids (lb/hr)':30}  {solids_in:>14,.1f}  {solids_out_sugar+solids_out_molasses:>14,.1f}  {solids_imbalance:>+14.4f}")
    print(f"    A sugar solids            : {solids_out_sugar:>14,.1f}")
    print(f"    Final molasses solids     : {solids_out_molasses:>14,.1f}")
    print()

    to_excel(result, filepath='boiling_scheme_output.xlsx')

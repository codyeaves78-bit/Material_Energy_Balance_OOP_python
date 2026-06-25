# Solve multiple EvaporatorSet instances simultaneously by balancing juice
# supply so that each active set reaches the same average U_calc/U_dessin ratio.
#
# Philosophy mirrors the old evaporator.py objective_function approach:
#   - n active sets need n-1 free variables (juice fractions)
#   - the last active set's fraction is derived as 1 - sum(others)
#   - scipy.optimize.fsolve zeroes the differences in mean U ratios between sets
#
# Inner loop (inside fsolve): solve_for_steam only — keeps pressures fixed,
#   fast and numerically stable even at off-nominal juice flows.
# Final solve: adjust_pressure_profile on each set to also balance pressures
#   within each set at the converged juice distribution.

from datetime import datetime

from EvaporatorSet import EvaporatorSet

_MIN_FRACTION = 0.02  # floor for any active set's juice fraction


def _mean_u_ratio(evap_set: EvaporatorSet) -> float:
    """Average U_calc/U_dessin across all effects in a set."""
    return sum(float(e.U_ratio) for e in evap_set.evaporator_list) / evap_set.number_of_effects


def _is_finite(value: float) -> bool:
    return value == value and value != float('inf') and value != float('-inf')


def solve_multi_set(
    evaporator_sets: list,
    online: list,
    total_juice_flow_lb_per_hr: float,
    vapor_demands: dict = None,
    bleed_toggles: list = None,
    set_names: list = None,
    excel_path: str = None,
    scenario_label: str = None,
    verbose: bool = True,
    pre_evaporator=None,
    bleed_overrides: dict = None,
) -> list:
    """
    Distribute total juice among multiple EvaporatorSet instances so that each
    active set achieves the same average U_calc/U_dessin ratio.

    Parameters
    ----------
    evaporator_sets : list[EvaporatorSet]
        All sets, online or not.  Each must already be constructed with its own
        juice_in SugarStream (brix / purity / temp already set); only the
        flow_lb_per_hr will be overwritten here.  Initial flow must be > 0.
    online : list[bool]
        Which sets are active.  Must be the same length as evaporator_sets.
    total_juice_flow_lb_per_hr : float
        Total juice flow (lb/hr) to split among the active sets.
    verbose : bool
        Print a summary of the final juice distribution when True.

    Returns
    -------
    list[float]
        Juice fraction assigned to each set (0.0 for offline sets).
        Fractions for active sets sum to 1.0.
    """
    num_total = len(evaporator_sets)
    active_indices = [i for i, on in enumerate(online) if on]
    num_active = len(active_indices)

    if num_active == 0:
        if verbose:
            print("No sets are online. Nothing to solve.")
        return [0.0] * num_total

    if set_names is None:
        set_names = [f"Set {i + 1}" for i in range(num_total)]
    resolved_toggles = _resolve_bleed_toggles(evaporator_sets, bleed_toggles)
    bleed_distribution: dict = {}
    if vapor_demands:
        bleed_distribution = compute_vapor_bleed_distribution(
            evaporator_sets, online, vapor_demands, resolved_toggles,
            verbose=verbose, bleed_overrides=bleed_overrides,
        )

    def _apply_full_solve(fractions: list) -> None:
        """Set juice flows and fully solve each active set (steam + pressure profile)."""
        for i in active_indices:
            es       = evaporator_sets[i]
            new_flow = fractions[i] * total_juice_flow_lb_per_hr
            old_flow = es.juice_in.flow_lb_per_hr
            # Scale steam initial guess proportionally to juice flow change.
            # Bleed correction has already been applied once (before the first call);
            # subsequent calls start from a converged steam value so plain scaling suffices.
            if old_flow > 0 and _is_finite(es.supply_steam.flow_lb_per_hr) and es.supply_steam.flow_lb_per_hr > 0:
                es.supply_steam.flow_lb_per_hr *= new_flow / old_flow
            es.juice_in.flow_lb_per_hr = new_flow
            es.adjust_pressure_profile()

    # ---- area-weighted starting fractions ----
    active_areas = [sum(evaporator_sets[i].effect_areas_ft2) for i in active_indices]
    total_active_area = sum(active_areas)
    final_fractions = [0.0] * num_total
    for idx, i in enumerate(active_indices):
        final_fractions[i] = active_areas[idx] / total_active_area

    # Pre-condition supply steam once, before the first full solve.
    # For each set: scale by the juice flow change AND add a bleed correction so
    # Newton-Raphson in solve_for_steam starts near the true solution.
    # Bleed at effect k starves (n-k-1) downstream effects, requiring extra
    # supply steam ≈ bleed_k * (n-k-1) / n (simple n-effect approximation).
    for i in active_indices:
        es = evaporator_sets[i]
        n  = es.number_of_effects
        bleed_correction = sum(
            es.evaporator_list[k].vapor_bleed.flow_lb_per_hr * (n - k - 1) / n
            for k in range(n - 1)
        )
        old_flow = es.juice_in.flow_lb_per_hr
        new_flow = final_fractions[i] * total_juice_flow_lb_per_hr
        if old_flow > 0 and _is_finite(es.supply_steam.flow_lb_per_hr) and es.supply_steam.flow_lb_per_hr > 0:
            base = es.supply_steam.flow_lb_per_hr * new_flow / old_flow
        else:
            base = es.steam_initial_guess
        es.supply_steam.flow_lb_per_hr = base + bleed_correction
        # Sync juice flow so _apply_full_solve sees old==new and doesn't re-scale steam.
        es.juice_in.flow_lb_per_hr = new_flow

    if num_active == 1:
        _apply_full_solve(final_fractions)
        if verbose:
            i = active_indices[0]
            print(f"\nSingle set online (Set {i + 1}): 100% of juice ({total_juice_flow_lb_per_hr:,.0f} lb/hr)")
            print(f"  Mean U ratio: {_mean_u_ratio(evaporator_sets[i]):.4f}")
        if excel_path:
            _export_to_excel(
                evaporator_sets, online, final_fractions, total_juice_flow_lb_per_hr,
                vapor_demands or {}, resolved_toggles, bleed_distribution, excel_path, set_names,
                scenario_label, pre=pre_evaporator,
            )
        return final_fractions

    # ---- iterative proportional control ----
    # Mirrors adjust_pressure_profile's dampened-ratio approach, applied between sets
    # instead of between effects.  Under-loaded sets (low U ratio) grow their fraction;
    # over-loaded sets shrink.  No external solver needed.
    max_iter   = 200
    outer_tol  = 0.005   # max U-ratio spread across active sets to declare convergence
    dampening  = 0.30    # fraction of proportional correction applied each step

    # Initial full solve at area-weighted fractions to establish pressure profiles
    _apply_full_solve(final_fractions)

    converged = False
    iteration = 0
    for iteration in range(max_iter):
        # Fast steam-only re-solve at current fractions (pressures held fixed)
        for i in active_indices:
            es       = evaporator_sets[i]
            new_flow = final_fractions[i] * total_juice_flow_lb_per_hr
            old_flow = es.juice_in.flow_lb_per_hr
            # Recover from NaN steam (can happen if a prior pressure adjustment failed);
            # also scale proportionally so the secant method doesn't diverge on flow changes.
            if not _is_finite(es.supply_steam.flow_lb_per_hr) or es.supply_steam.flow_lb_per_hr <= 0:
                es.supply_steam.flow_lb_per_hr = es.steam_initial_guess * (new_flow / (total_juice_flow_lb_per_hr / num_total))
            elif old_flow > 0:
                es.supply_steam.flow_lb_per_hr *= new_flow / old_flow
            es.juice_in.flow_lb_per_hr = new_flow
            es.solve_for_steam()

        u_vals = [_mean_u_ratio(evaporator_sets[i]) for i in active_indices]

        if not all(_is_finite(u) for u in u_vals):
            if verbose:
                print(f"  Iteration {iteration + 1}: non-finite U ratio — stopping early.")
            break

        spread = max(u_vals) - min(u_vals)
        if spread <= outer_tol:
            converged = True
            break

        # Proportional correction: ratio > 1 means set is under-loaded → gets more juice
        avg_u = sum(u_vals) / num_active
        prev_fractions = [final_fractions[i] for i in active_indices]
        for idx, i in enumerate(active_indices):
            ratio = avg_u / u_vals[idx] if u_vals[idx] != 0 else 1.0
            if ratio > 0:   # negative ratio means non-physical U; skip correction
                final_fractions[i] *= ratio ** dampening

        # Clamp floor then renormalise
        for i in active_indices:
            final_fractions[i] = max(_MIN_FRACTION, final_fractions[i])
        s = sum(final_fractions[i] for i in active_indices)
        for i in active_indices:
            final_fractions[i] /= s

        # Secondary convergence: fractions have stopped moving (physical equilibrium
        # reached — e.g. sets with different effect counts naturally settle at different
        # U ratios).  Treat as converged even if spread > outer_tol.
        max_frac_change = max(
            abs(final_fractions[active_indices[j]] - prev_fractions[j])
            for j in range(num_active)
        )
        if max_frac_change < 1e-5:
            converged = True
            break

    # Final full solve at converged fractions — polishes pressures and U ratios
    _apply_full_solve(final_fractions)

    if converged:
        if verbose:
            print(f"\nJuice distribution converged in {iteration + 1} iteration(s).")
            _print_distribution_summary(evaporator_sets, online, final_fractions, total_juice_flow_lb_per_hr)
    else:
        print(f"\nWARNING: Did not converge after {max_iter} iterations — reporting best estimate.")
        if verbose:
            _print_distribution_summary(evaporator_sets, online, final_fractions, total_juice_flow_lb_per_hr)

    if excel_path:
        _export_to_excel(
            evaporator_sets, online, final_fractions, total_juice_flow_lb_per_hr,
            vapor_demands or {}, resolved_toggles, bleed_distribution, excel_path, set_names,
            scenario_label, pre=pre_evaporator,
        )

    return final_fractions


def _print_distribution_summary(
    evaporator_sets: list,
    online: list,
    fractions: list,
    total_juice_flow_lb_per_hr: float
) -> None:
    """Print a concise summary of the juice distribution result."""
    print(f"\n{'='*60}")
    print(f"  Multi-Set Juice Distribution Summary")
    print(f"  Total juice: {total_juice_flow_lb_per_hr:,.0f} lb/hr")
    print(f"{'='*60}")
    for i, (evap_set, on, frac) in enumerate(zip(evaporator_sets, online, fractions)):
        status = "ONLINE" if on else "OFFLINE"
        if on:
            flow = frac * total_juice_flow_lb_per_hr
            u = _mean_u_ratio(evap_set)
            steam = evap_set.supply_steam.flow_lb_per_hr
            brix_out = evap_set.evaporator_list[-1].juice_side_out.brix
            print(
                f"  Set {i + 1} [{status}]: {frac*100:.2f}%  "
                f"({flow:,.0f} lb/hr juice)  "
                f"Steam: {steam:,.0f} lb/hr  "
                f"Syrup brix: {brix_out:.2f}  "
                f"Mean U ratio: {u:.4f}"
            )
        else:
            print(f"  Set {i + 1} [{status}]: 0%")
    print(f"{'='*60}\n")


# ---------------------------------------------------------------------------
# Vapor bleed distribution helpers
# ---------------------------------------------------------------------------

def _resolve_bleed_toggles(evaporator_sets: list, bleed_toggles: list) -> list:
    """Return bleed_toggles, substituting all-True for any set whose entry is None or missing."""
    resolved = []
    for i, evap_set in enumerate(evaporator_sets):
        n_bleedable = evap_set.number_of_effects - 1
        if bleed_toggles is None or i >= len(bleed_toggles) or bleed_toggles[i] is None:
            resolved.append([True] * n_bleedable)
        else:
            resolved.append(list(bleed_toggles[i]))
    return resolved


def compute_vapor_bleed_distribution(
    evaporator_sets: list,
    online: list,
    vapor_demands: dict,
    bleed_toggles: list,
    verbose: bool = True,
    bleed_overrides: dict = None,
) -> dict:
    """
    Distribute total vapor bleed demands across active sets by heating-surface share.

    Updates each active set's vapor_bleeds list *and* the live
    evaporator_list[k].vapor_bleed.flow_lb_per_hr in-place so subsequent
    solve calls see the new values.

    Parameters
    ----------
    vapor_demands : dict  {effect_number (1-based): total_lb_per_hr}
        e.g. {1: 500_000, 2: 200_000}
    bleed_toggles : list[list[bool]]
        bleed_toggles[i][k] — True if set i participates in effect-(k+1) bleeding.
        Use _resolve_bleed_toggles() to fill in defaults before calling here.
    bleed_overrides : dict, optional  {effect_number: {set_index: lb_per_hr}}
        Per-set bleed allocations that bypass area-weighted distribution for
        that effect.  Used during U-ratio-based V1 redistribution.

    Returns
    -------
    dict  {effect_number: {set_index: lb_per_hr}}
    """
    num_total = len(evaporator_sets)

    # Pad vapor_bleeds on every set so in-place writes are safe
    for evap_set in evaporator_sets:
        n_bleedable = evap_set.number_of_effects - 1
        while len(evap_set.vapor_bleeds) < n_bleedable:
            evap_set.vapor_bleeds.append(0.0)

    distribution = {}

    for effect_num, total_demand in sorted(vapor_demands.items()):
        k = effect_num - 1  # 0-indexed into vapor_bleeds / effect_areas_ft2
        distribution[effect_num] = {i: 0.0 for i in range(num_total)}

        if bleed_overrides and effect_num in bleed_overrides:
            # Use caller-supplied per-set allocations (U-ratio redistribution)
            for i, flow in bleed_overrides[effect_num].items():
                distribution[effect_num][i] = flow
        else:
            participating = [
                i for i, (es, on) in enumerate(zip(evaporator_sets, online))
                if on
                and k < es.number_of_effects - 1
                and k < len(bleed_toggles[i])
                and bleed_toggles[i][k]
            ]

            if not participating:
                if total_demand > 0 and verbose:
                    print(
                        f"WARNING: no active sets can supply V{effect_num} "
                        f"demand of {total_demand:,.0f} lb/hr"
                    )
                continue

            total_area = sum(evaporator_sets[i].effect_areas_ft2[k] for i in participating)
            if total_area <= 0:
                if verbose:
                    print(f"WARNING: total area for V{effect_num} is zero; bleed unallocated.")
                continue

            for i in participating:
                share = evaporator_sets[i].effect_areas_ft2[k] / total_area
                distribution[effect_num][i] = total_demand * share

    # Write back to each set's vapor_bleeds list AND to the live Evaporator objects.
    # Also warn if any assigned bleed exceeds that effect's current vapor production —
    # a bleed > lbs_evaporated makes vapor_to_next go negative and U ratios blow up.
    for i, (evap_set, on) in enumerate(zip(evaporator_sets, online)):
        if not on:
            continue
        for k in range(evap_set.number_of_effects - 1):
            effect_num = k + 1
            if effect_num in distribution:
                flow = distribution[effect_num].get(i, 0.0)
                evap_set.vapor_bleeds[k] = flow
                evap_set.evaporator_list[k].vapor_bleed.flow_lb_per_hr = flow
                available = evap_set.evaporator_list[k].lbs_evaporated_per_hr
                if flow > available and verbose:
                    print(
                        f"WARNING: Set {i + 1} effect {effect_num} bleed "
                        f"{flow:,.0f} lb/hr exceeds current vapor production "
                        f"{available:,.0f} lb/hr — reduce V{effect_num} demand "
                        f"or the solver will see negative steam in downstream effects."
                    )

    return distribution


def _export_to_excel(
    evaporator_sets: list,
    online: list,
    fractions: list,
    total_juice_flow_lb_per_hr: float,
    vapor_demands: dict,
    bleed_toggles: list,
    bleed_distribution: dict,
    path: str,
    set_names: list,
    scenario_label: str = None,
    pre=None,
) -> None:
    """Write an Excel workbook: distribution summary, per-effect engineering detail,
    vapor bleed allocation, and an embedded PFD diagram for every online set.
    If scenario_label is provided it appears in the report title row.
    If pre (a PreEvaporator) is provided, a Pre detail sheet and PFD are added."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image as XLImage
    except ImportError:
        print("WARNING: openpyxl not installed — skipping Excel export.  pip install openpyxl")
        return

    # ---- style helpers ------------------------------------------------
    _DARK   = "2E5F8A"
    _MID    = "4A90C4"
    _LIGHT  = "D9E8F5"
    _ALT    = "F2F7FC"
    _WHITE  = "FFFFFF"

    def _fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def _bd():
        s = Side(style="thin", color="BBBBBB")
        return Border(left=s, right=s, top=s, bottom=s)

    def _al(h="right", wrap=False):
        return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

    def title_cell(c, text, size=14):
        c.value = text
        c.font  = Font(bold=True, size=size, color=_DARK)
        c.alignment = _al("center")

    def hdr(c, text):
        c.value = text
        c.font  = Font(bold=True, color=_WHITE)
        c.fill  = _fill(_DARK)
        c.border = _bd()
        c.alignment = _al("center", wrap=True)

    def sub_hdr(c, text):
        c.value = text
        c.font  = Font(bold=True, color=_WHITE)
        c.fill  = _fill(_MID)
        c.border = _bd()
        c.alignment = _al("left")

    def dat_alt(c, val, fmt=None, row_idx=0, center=False, left=False):
        c.value = val
        c.font  = Font()
        c.border = _bd()
        c.fill  = _fill(_ALT) if row_idx % 2 == 0 else _fill(_WHITE)
        c.alignment = _al("center" if center else ("left" if left else "right"))
        if fmt:
            c.number_format = fmt

    def tot(c, val, fmt=None, left=False):
        c.value = val
        c.font  = Font(bold=True)
        c.fill  = _fill(_LIGHT)
        c.border = _bd()
        c.alignment = _al("left" if left else "right")
        if fmt:
            c.number_format = fmt

    def col_w(ws, widths):
        for idx, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = w

    def _vac_fmt(psia: float) -> str:
        """psig above atmosphere; inches-Hg vacuum below."""
        if psia >= 14.696:
            return f"{psia - 14.696:.2f} psig"
        return f'{29.92 - 29.92 * psia / 14.696:.2f}" Hg Vac'

    stamp = datetime.now().strftime("%Y-%m-%d  %H:%M")
    wb = openpyxl.Workbook()

    # ==================================================================
    # Sheet 1 — Distribution Summary
    # ==================================================================
    ws1 = wb.active
    ws1.title = "Distribution Summary"
    ws1.sheet_view.showGridLines = False

    ws1.merge_cells("A1:J1")
    _report_title = "Multi-Set Evaporator  —  Distribution Summary"
    if scenario_label:
        _report_title = f"{scenario_label}  |  {_report_title}"
    title_cell(ws1["A1"], _report_title)
    ws1["A2"] = (
        f"Total juice feed: {total_juice_flow_lb_per_hr:,.0f} lb/hr      Generated: {stamp}"
    )
    ws1["A2"].font = Font(italic=True, color="555555")

    S1_HDRS = [
        "Set", "Status", "# Effects", "Juice Flow\n(lb/hr)", "Fraction\n(%)",
        "Supply Steam\n(lb/hr)", "Syrup Brix\n(°Brix)", "Water Evap'd\n(lb/hr)",
        "Steam Economy\n(lb water/lb steam)", "Mean U Ratio",
    ]
    for col, h in enumerate(S1_HDRS, 1):
        hdr(ws1.cell(row=4, column=col), h)
    ws1.row_dimensions[4].height = 36

    r = 5
    tot_juice = 0.0
    tot_steam = 0.0
    tot_water = 0.0

    if pre is not None:
        pre_econ = (pre.vapor_bleed_lb_per_hr / pre.exhaust_required_lb_per_hr
                    if pre.exhaust_required_lb_per_hr > 0 else 0.0)
        dat_alt(ws1.cell(r, 1),  "Pre Evaporator",               row_idx=0, left=True)
        dat_alt(ws1.cell(r, 2),  "ONLINE",                       row_idx=0, center=True)
        dat_alt(ws1.cell(r, 3),  1,                              row_idx=0, center=True)
        dat_alt(ws1.cell(r, 4),  pre.juice_in.flow_lb_per_hr,    "#,##0",  row_idx=0)
        dat_alt(ws1.cell(r, 5),  "—",                            row_idx=0, center=True)
        dat_alt(ws1.cell(r, 6),  pre.exhaust_required_lb_per_hr, "#,##0",  row_idx=0)
        dat_alt(ws1.cell(r, 7),  pre.juice_out.brix,             "0.00",   row_idx=0)
        dat_alt(ws1.cell(r, 8),  pre.vapor_bleed_lb_per_hr,      "#,##0",  row_idx=0)
        dat_alt(ws1.cell(r, 9),  pre_econ,                       "0.000",  row_idx=0)
        dat_alt(ws1.cell(r, 10), "—",                            row_idx=0, center=True)
        r += 1

    for i, (es, on, frac) in enumerate(zip(evaporator_sets, online, fractions)):
        ri = r - 5  # 0-based for alternating
        dat_alt(ws1.cell(r, 1), set_names[i],       row_idx=ri, left=True)
        dat_alt(ws1.cell(r, 2), "ONLINE" if on else "OFFLINE", row_idx=ri, center=True)
        dat_alt(ws1.cell(r, 3), es.number_of_effects if on else "—",
                row_idx=ri, center=True)
        if on:
            flow   = frac * total_juice_flow_lb_per_hr
            steam  = es.supply_steam.flow_lb_per_hr
            brix   = es.evaporator_list[-1].juice_side_out.brix
            water  = sum(ev.lbs_evaporated_per_hr for ev in es.evaporator_list)
            econ   = (water / steam) if steam > 0 else 0.0
            u_mean = _mean_u_ratio(es)
            tot_juice += flow
            tot_steam += steam
            tot_water += water
            dat_alt(ws1.cell(r, 4),  flow,   "#,##0",  row_idx=ri)
            dat_alt(ws1.cell(r, 5),  frac,   "0.0%",   row_idx=ri)
            dat_alt(ws1.cell(r, 6),  steam,  "#,##0",  row_idx=ri)
            dat_alt(ws1.cell(r, 7),  brix,   "0.00",   row_idx=ri)
            dat_alt(ws1.cell(r, 8),  water,  "#,##0",  row_idx=ri)
            dat_alt(ws1.cell(r, 9),  econ,   "0.000",  row_idx=ri)
            dat_alt(ws1.cell(r, 10), u_mean, "0.0000", row_idx=ri)
        else:
            for col in range(4, 11):
                dat_alt(ws1.cell(r, col), "—", row_idx=ri, center=True)
        r += 1

    for col in range(1, 11):
        tot(ws1.cell(r, col), None)
    tot(ws1.cell(r, 1), "TOTAL",     left=True)
    tot(ws1.cell(r, 4), tot_juice,   "#,##0")
    tot(ws1.cell(r, 5), 1.0,         "0.0%")
    tot(ws1.cell(r, 6), tot_steam,   "#,##0")
    tot(ws1.cell(r, 8), tot_water,   "#,##0")
    tot(ws1.cell(r, 9), (tot_water / tot_steam) if tot_steam > 0 else 0.0, "0.000")

    col_w(ws1, [16, 10, 10, 18, 12, 18, 14, 18, 22, 14])

    # ==================================================================
    # Sheet 2 — Effect Details
    # ==================================================================
    ws2 = wb.create_sheet(title="Effect Details")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:U1")
    title_cell(ws2["A1"], "Per-Effect Engineering Detail")
    ws2["A2"] = f"Generated: {stamp}"
    ws2["A2"].font = Font(italic=True, color="555555")

    EFF_HDRS = [
        "Effect #", "Area\n(ft²)",
        "Steam/Vapor In\n(lb/hr)", "Steam P\n(psia)", "Steam T\n(°F)",
        "ΔT\n(°F)",
        "Juice In\n(lb/hr)", "Juice In\n(°Brix)", "Juice In\n(°F)",
        "Juice Out\n(lb/hr)", "Juice Out\n(°Brix)", "Juice Out\n(°F)",
        "BPE\n(°F)", "Vapor Out\n(lb/hr)", "Vapor Bleed\n(lb/hr)",
        "Vapor P\n(psia)", "Vapor P\n(psig / \" Hg Vac)", "Vapor T\n(°F)",
        "U calc\n(BTU/hr·ft²·°F)", "U Dessin\n(BTU/hr·ft²·°F)", "U Ratio",
    ]
    N_EFF_COLS = len(EFF_HDRS)  # 21

    cur = 4

    if pre is not None:
        ws2.merge_cells(f"A{cur}:U{cur}")
        sub_hdr(ws2.cell(cur, 1),
            f"Pre Evaporator  —  ONLINE  |  "
            f"Juice In: {pre.juice_in.flow_lb_per_hr:,.0f} lb/hr  |  Effects: 1"
        )
        ws2.row_dimensions[cur].height = 20
        cur += 1

        for col, h in enumerate(EFF_HDRS, 1):
            hdr(ws2.cell(cur, col), h)
        ws2.row_dimensions[cur].height = 36
        cur += 1

        bpe_pre = pre.liquid_temp_deg_F - pre.vapor_temp_deg_F
        dt_pre  = pre.supply_steam.sat_temp_deg_F - pre.liquid_temp_deg_F
        dat_alt(ws2.cell(cur, 1),  1,                                row_idx=0, center=True)
        dat_alt(ws2.cell(cur, 2),  pre.area_ft2,                     "#,##0",  row_idx=0)
        dat_alt(ws2.cell(cur, 3),  pre.exhaust_required_lb_per_hr,   "#,##0",  row_idx=0)
        dat_alt(ws2.cell(cur, 4),  pre.supply_steam.P_psia,          "0.00",   row_idx=0)
        dat_alt(ws2.cell(cur, 5),  pre.supply_steam.sat_temp_deg_F,  "0.0",    row_idx=0)
        dat_alt(ws2.cell(cur, 6),  dt_pre,                           "0.0",    row_idx=0)
        dat_alt(ws2.cell(cur, 7),  pre.juice_in.flow_lb_per_hr,      "#,##0",  row_idx=0)
        dat_alt(ws2.cell(cur, 8),  pre.juice_in.brix,                "0.00",   row_idx=0)
        dat_alt(ws2.cell(cur, 9),  pre.juice_in.temp_deg_F,          "0.0",    row_idx=0)
        dat_alt(ws2.cell(cur, 10), pre.juice_out.flow_lb_per_hr,     "#,##0",  row_idx=0)
        dat_alt(ws2.cell(cur, 11), pre.juice_out.brix,               "0.00",   row_idx=0)
        dat_alt(ws2.cell(cur, 12), pre.juice_out.temp_deg_F,         "0.0",    row_idx=0)
        dat_alt(ws2.cell(cur, 13), bpe_pre,                          "0.00",   row_idx=0)
        dat_alt(ws2.cell(cur, 14), pre.vapor_bleed_lb_per_hr,        "#,##0",  row_idx=0)
        dat_alt(ws2.cell(cur, 15), pre.vapor_bleed_lb_per_hr,        "#,##0",  row_idx=0)
        dat_alt(ws2.cell(cur, 16), pre.vapor_pressure_psia,          "0.00",   row_idx=0)
        dat_alt(ws2.cell(cur, 17), _vac_fmt(pre.vapor_pressure_psia), row_idx=0, left=True)
        dat_alt(ws2.cell(cur, 18), pre.vapor_temp_deg_F,             "0.0",    row_idx=0)
        dat_alt(ws2.cell(cur, 19), "—",                              row_idx=0, center=True)
        dat_alt(ws2.cell(cur, 20), pre.dessin_U,                     "0.0",    row_idx=0)
        dat_alt(ws2.cell(cur, 21), "—",                              row_idx=0, center=True)
        cur += 2  # blank row after Pre section

    for i, (es, on, frac) in enumerate(zip(evaporator_sets, online, fractions)):
        # Set banner spanning all columns
        ws2.merge_cells(f"A{cur}:U{cur}")
        flow_str = f"{frac * total_juice_flow_lb_per_hr:,.0f} lb/hr" if on else "OFFLINE"
        sub_hdr(
            ws2.cell(cur, 1),
            f"{set_names[i]}  —  {'ONLINE' if on else 'OFFLINE'}  |  "
            f"Juice: {flow_str}  |  Effects: {es.number_of_effects}",
        )
        ws2.row_dimensions[cur].height = 20
        cur += 1

        if not on:
            cur += 1
            continue

        # Column headers
        for col, h in enumerate(EFF_HDRS, 1):
            hdr(ws2.cell(cur, col), h)
        ws2.row_dimensions[cur].height = 36
        cur += 1

        # Per-effect rows
        set_area   = 0.0
        set_steam  = 0.0
        set_water  = 0.0
        set_bleed  = 0.0
        ri = 0
        for k, evap in enumerate(es.evaporator_list):
            c1  = evap.calandria_side
            jin = evap.juice_side_in
            jout = evap.juice_side_out
            dt  = evap.delta_T_juice_steam
            dat_alt(ws2.cell(cur, 1),  k + 1,                    row_idx=ri, center=True)
            dat_alt(ws2.cell(cur, 2),  evap.area_ft2,            "#,##0",  row_idx=ri)
            dat_alt(ws2.cell(cur, 3),  c1.flow_lb_per_hr,        "#,##0",  row_idx=ri)
            dat_alt(ws2.cell(cur, 4),  c1.P_psia,                "0.00",   row_idx=ri)
            dat_alt(ws2.cell(cur, 5),  c1.sat_temp_deg_F,        "0.0",    row_idx=ri)
            dat_alt(ws2.cell(cur, 6),  dt,                       "0.0",    row_idx=ri)
            dat_alt(ws2.cell(cur, 7),  jin.flow_lb_per_hr,       "#,##0",  row_idx=ri)
            dat_alt(ws2.cell(cur, 8),  jin.brix,                 "0.00",   row_idx=ri)
            dat_alt(ws2.cell(cur, 9),  jin.temp_deg_F,           "0.0",    row_idx=ri)
            dat_alt(ws2.cell(cur, 10), jout.flow_lb_per_hr,      "#,##0",  row_idx=ri)
            dat_alt(ws2.cell(cur, 11), jout.brix,                "0.00",   row_idx=ri)
            dat_alt(ws2.cell(cur, 12), jout.temp_deg_F,          "0.0",    row_idx=ri)
            dat_alt(ws2.cell(cur, 13), evap.bpe_juice,           "0.00",   row_idx=ri)
            dat_alt(ws2.cell(cur, 14), evap.vapor_out.flow_lb_per_hr,   "#,##0",  row_idx=ri)
            dat_alt(ws2.cell(cur, 15), evap.vapor_bleed.flow_lb_per_hr, "#,##0",  row_idx=ri)
            dat_alt(ws2.cell(cur, 16), evap.vapor_pressure_psia,        "0.00",   row_idx=ri)
            dat_alt(ws2.cell(cur, 17), _vac_fmt(evap.vapor_pressure_psia), row_idx=ri, left=True)
            dat_alt(ws2.cell(cur, 18), evap.vapor_temperature,          "0.0",    row_idx=ri)
            dat_alt(ws2.cell(cur, 19), evap.heat_xfer_U,                "0.0",    row_idx=ri)
            dat_alt(ws2.cell(cur, 20), evap.dessin_U,                   "0.0",    row_idx=ri)
            dat_alt(ws2.cell(cur, 21), evap.U_ratio,                    "0.0000", row_idx=ri)
            set_area  += evap.area_ft2
            set_steam += c1.flow_lb_per_hr
            set_water += evap.lbs_evaporated_per_hr
            set_bleed += evap.vapor_bleed.flow_lb_per_hr
            cur += 1
            ri  += 1

        # Set totals row
        for col in range(1, N_EFF_COLS + 1):
            tot(ws2.cell(cur, col), None)
        tot(ws2.cell(cur, 1),  f"{set_names[i]} Total", left=True)
        tot(ws2.cell(cur, 2),  set_area,  "#,##0")
        tot(ws2.cell(cur, 3),  set_steam, "#,##0")
        tot(ws2.cell(cur, 14), set_water, "#,##0")
        tot(ws2.cell(cur, 15), set_bleed, "#,##0")
        cur += 2  # blank row between sets

    col_w(ws2, [10, 10, 18, 12, 12, 10, 15, 13, 12, 15, 13, 13, 10, 15, 15, 12, 18, 12, 18, 18, 12])

    # ==================================================================
    # Sheet 3 — Vapor Bleed Distribution
    # ==================================================================
    ws3 = wb.create_sheet(title="Vapor Bleeds")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells("A1:F1")
    title_cell(ws3["A1"], "Vapor Bleed Distribution by Effect")
    ws3["A2"] = f"Generated: {stamp}"
    ws3["A2"].font = Font(italic=True, color="555555")

    cur = 4

    if not vapor_demands:
        ws3.cell(cur, 1).value = "No vapor demands specified."
    else:
        BLEED_HDRS = [
            "Set", "Status", "Effect Area\n(ft²)", "Contributing?",
            "Area Share\n(%)", "Bleed Flow\n(lb/hr)",
        ]

        for effect_num in sorted(vapor_demands.keys()):
            total_demand = vapor_demands[effect_num]
            k = effect_num - 1  # 0-indexed

            # Pre contributes V1 only; reconstruct true total demand for the banner
            pre_contributes = (pre is not None and effect_num == 1)
            pre_area_v1     = pre.area_ft2 if pre_contributes else 0.0
            display_demand  = total_demand + (pre.vapor_bleed_lb_per_hr if pre_contributes else 0.0)

            ws3.merge_cells(f"A{cur}:F{cur}")
            sub_hdr(
                ws3.cell(cur, 1),
                f"V{effect_num}  —  Effect {effect_num} Vapor     "
                f"Total Demand: {display_demand:,.0f} lb/hr",
            )
            cur += 1

            for col, h in enumerate(BLEED_HDRS, 1):
                hdr(ws3.cell(cur, col), h)
            ws3.row_dimensions[cur].height = 30
            cur += 1

            participating = {
                i for i, (es, on) in enumerate(zip(evaporator_sets, online))
                if on
                and k < es.number_of_effects - 1
                and k < len(bleed_toggles[i])
                and bleed_toggles[i][k]
            }
            part_area_sum = (
                sum(evaporator_sets[i].effect_areas_ft2[k] for i in participating)
                if participating else 0.0
            )
            total_area = part_area_sum + pre_area_v1  # used for all share calculations

            sum_area  = 0.0
            sum_bleed = 0.0
            ri = 0

            # Pre Evaporator row (V1 only)
            if pre_contributes:
                pre_share = pre_area_v1 / total_area if total_area > 0 else 0.0
                dat_alt(ws3.cell(cur, 1), "Pre Evaporator",              row_idx=ri, left=True)
                dat_alt(ws3.cell(cur, 2), "ONLINE",                      row_idx=ri, center=True)
                dat_alt(ws3.cell(cur, 3), pre_area_v1,        "#,##0",   row_idx=ri)
                dat_alt(ws3.cell(cur, 4), "Yes",                         row_idx=ri, center=True)
                dat_alt(ws3.cell(cur, 5), pre_share,          "0.0%",    row_idx=ri)
                dat_alt(ws3.cell(cur, 6), pre.vapor_bleed_lb_per_hr, "#,##0", row_idx=ri)
                sum_area  += pre_area_v1
                sum_bleed += pre.vapor_bleed_lb_per_hr
                cur += 1
                ri  += 1

            for i, (es, on) in enumerate(zip(evaporator_sets, online)):
                dat_alt(ws3.cell(cur, 1), set_names[i], row_idx=ri, left=True)
                dat_alt(ws3.cell(cur, 2), "ONLINE" if on else "OFFLINE",
                        row_idx=ri, center=True)

                has_eff = on and k < es.number_of_effects - 1
                if has_eff:
                    area       = es.effect_areas_ft2[k]
                    contrib    = i in participating
                    bleed_flow = bleed_distribution.get(effect_num, {}).get(i, 0.0)
                    share      = (area / total_area) if (contrib and total_area > 0) else 0.0
                    dat_alt(ws3.cell(cur, 3), area,       "#,##0", row_idx=ri)
                    dat_alt(ws3.cell(cur, 4), "Yes" if contrib else "No",
                            row_idx=ri, center=True)
                    dat_alt(ws3.cell(cur, 5), share,      "0.0%",  row_idx=ri)
                    dat_alt(ws3.cell(cur, 6), bleed_flow, "#,##0", row_idx=ri)
                    if contrib:
                        sum_area  += area
                        sum_bleed += bleed_flow
                else:
                    note = "—" if not on else f"No eff {effect_num}"
                    for col in range(3, 7):
                        dat_alt(ws3.cell(cur, col), note, row_idx=ri, center=True)
                cur += 1
                ri  += 1

            for col in range(1, 7):
                tot(ws3.cell(cur, col), None)
            tot(ws3.cell(cur, 1), "TOTAL",    left=True)
            tot(ws3.cell(cur, 3), sum_area,   "#,##0")
            tot(ws3.cell(cur, 5), 1.0 if sum_bleed > 0 else 0.0, "0.0%")
            tot(ws3.cell(cur, 6), sum_bleed,  "#,##0")
            cur += 2

    col_w(ws3, [16, 10, 18, 14, 16, 22])

    # ==================================================================
    # Pre Evaporator detail sheet (no matplotlib dependency)
    # ==================================================================
    if pre is not None:
        ws_pre = wb.create_sheet(title="Pre Evaporator")
        ws_pre.sheet_view.showGridLines = False
        ws_pre.merge_cells("A1:C1")
        title_cell(ws_pre["A1"], "Pre Evaporator  —  Engineering Detail", size=13)
        ws_pre["A2"] = f"Status: ONLINE      Generated: {stamp}"
        ws_pre["A2"].font = Font(italic=True, color="555555")

        PRE_ROWS = [
            ("JUICE IN",            None,  None),
            ("  Flow",              pre.juice_in.flow_lb_per_hr,       "lb/hr"),
            ("  Brix",              pre.juice_in.brix,                  "°Brix"),
            ("  Purity",            pre.juice_in.purity,                "%"),
            ("  Temperature",       pre.juice_in.temp_deg_F,            "°F"),
            ("SUPPLY STEAM",        None,  None),
            ("  Pressure",          pre.supply_steam.P_psia,            "psia"),
            ("  Sat Temp",          pre.supply_steam.sat_temp_deg_F,    "°F"),
            ("  Flow Used",         pre.exhaust_required_lb_per_hr,     "lb/hr"),
            ("V1 VAPOR BLEED",      None,  None),
            ("  Flow",              pre.vapor_bleed_lb_per_hr,          "lb/hr"),
            ("  Pressure (psia)",   pre.vapor_pressure_psia,            "psia"),
            ("  Pressure (2nd)",    _vac_fmt(pre.vapor_pressure_psia),  ""),
            ("  Temperature",       pre.vapor_temp_deg_F,               "°F"),
            ("JUICE OUT",           None,  None),
            ("  Flow",              pre.juice_out.flow_lb_per_hr,       "lb/hr"),
            ("  Brix",              pre.juice_out.brix,                 "°Brix"),
            ("  Temperature",       pre.juice_out.temp_deg_F,           "°F"),
            ("PERFORMANCE",         None,  None),
            ("  Heat Duty",         pre.heat_duty_btu_per_hr,           "BTU/hr"),
            ("  U Dessin",          pre.dessin_U,                       "BTU/hr·ft²·°F"),
            ("  Heating Area",      pre.area_ft2,                       "ft²"),
            ("  Liquid Level",      pre.liquid_level_ft,                "ft"),
        ]

        for col, hdr_txt in enumerate(["Parameter", "Value", "Units"], 1):
            hdr(ws_pre.cell(4, col), hdr_txt)

        for ri, (param, val, unit) in enumerate(PRE_ROWS):
            row = 5 + ri
            is_header = val is None
            c_param = ws_pre.cell(row, 1)
            c_val   = ws_pre.cell(row, 2)
            c_unit  = ws_pre.cell(row, 3)
            if is_header:
                sub_hdr(c_param, param)
                for c in (c_val, c_unit):
                    c.fill   = _fill(_MID)
                    c.border = _bd()
            else:
                fmt = "#,##0" if isinstance(val, float) and val > 100 else "0.00##"
                dat_alt(c_param, param, row_idx=ri, left=True)
                dat_alt(c_val,   val,   fmt if isinstance(val, float) else None, row_idx=ri,
                        left=isinstance(val, str))
                dat_alt(c_unit,  unit,  row_idx=ri, left=True)

        ws_pre.column_dimensions["A"].width = 24
        ws_pre.column_dimensions["B"].width = 18
        ws_pre.column_dimensions["C"].width = 20

    # ==================================================================
    # PFD diagram sheets — one per online set + Pre (matplotlib optional)
    # ==================================================================
    try:
        import io as _io
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as _plt
        from evaporator_diagram import plot_set_diagram as _plot_pfd

        for i, (es, on) in enumerate(zip(evaporator_sets, online)):
            if not on:
                continue
            sname = set_names[i]
            sheet_name = (sname[:26].strip() + " PFD")[:31]
            ws_d = wb.create_sheet(title=sheet_name)
            ws_d.sheet_view.showGridLines = False
            ws_d.merge_cells("A1:M1")
            title_cell(ws_d["A1"], f"{sname}  —  Process Flow Diagram", size=12)

            try:
                fig = _plot_pfd(es, set_name=sname, show=False)
                buf = _io.BytesIO()
                fig.savefig(buf, format="png", dpi=110, bbox_inches="tight")
                buf.seek(0)
                img = XLImage(buf)
                img.width  = 1000
                img.height = 760
                ws_d.add_image(img, "A2")
                _plt.close(fig)
            except Exception as exc:
                ws_d.cell(3, 1).value = f"Diagram not available: {exc}"

        if pre is not None:
            from evaporator_diagram import plot_pre_diagram as _plot_pre_pfd
            ws_ppfd = wb.create_sheet(title="Pre Evaporator PFD")
            ws_ppfd.sheet_view.showGridLines = False
            ws_ppfd.merge_cells("A1:M1")
            title_cell(ws_ppfd["A1"], "Pre Evaporator  —  Process Flow Diagram", size=12)
            try:
                fig = _plot_pre_pfd(pre, pre_name="Pre Evaporator", show=False)
                buf = _io.BytesIO()
                fig.savefig(buf, format="png", dpi=110, bbox_inches="tight")
                buf.seek(0)
                img = XLImage(buf)
                img.width  = 1000
                img.height = 760
                ws_ppfd.add_image(img, "A2")
                _plt.close(fig)
            except Exception as exc:
                ws_ppfd.cell(3, 1).value = f"Diagram not available: {exc}"

    except Exception as exc:
        print(f"WARNING: PFD diagram export skipped — {exc}")

    wb.save(path)
    print(f"\nExcel report saved -> {path}")


# ---------------------------------------------------------------------------
# Stress test — run directly:  python multi_set_solver.py
# Tests 1-set / 2-set / 3-set / 4-set configurations with bleeds ON / OFF / mixed.
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import os
    import re
    import time
    from SugarStream import SugarStream
    from SteamStream import EvaporatorSteam
    from evaporator_functions import convert_inHg_vacuum_to_psia, convert_psig_to_psia
    from PreEvaporator import PreEvaporator

    _HERE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Excel File Outputs")

    total_juice = 1_500_000   # lb/hr  (~750 tph clear juice)
    exh_psia    = convert_psig_to_psia(15)
    last_psia   = convert_inHg_vacuum_to_psia(25)
    prim_heater_demand = 90_000
    sec_heater_demand = 90_000
    pan_demand = 180_000

    demands     = {1: pan_demand + sec_heater_demand, 2: prim_heater_demand}   # lb/hr total V1 / V2 demand

    _PRE_AREA     = 35_000
    _SET_V1_AREAS = [25000, 12000, 11000, 15000]   # first-effect area of each set

    NAMES = [
        "Set 1 (4-eff 25k ft²)",
        "Set 2 (4-eff 12k ft²)",
        "Set 3 (3-eff 11-9k ft²)",
        "Set 4 (4-eff 15k ft²)",
    ]

    def make_sets(juice_brix=14, juice_temp_deg_F=225, juice_flow_per_set=None):
        """Return 4 freshly-constructed EvaporatorSet instances."""
        if juice_flow_per_set is None:
            juice_flow_per_set = total_juice / 4
        def juice():
            return SugarStream(
                brix=juice_brix, purity=90, flow_lb_per_hr=juice_flow_per_set,
                temp_deg_F=juice_temp_deg_F, pressure_psia=60, level_ft=0,
            )
        def steam():
            return EvaporatorSteam(P_psia=exh_psia, flow_lb_per_hr=0)

        return [
            EvaporatorSet(
                juice_in=juice(), supply_steam=steam(),
                last_effect_pressure_psia=last_psia, target_brix_out=65,
                effect_areas_ft2=[25000, 25000, 25000, 25000],
                vapor_bleeds=[0, 0, 0], dessin_coefficient=18000, liquid_level_ft=2,
            ),
            EvaporatorSet(
                juice_in=juice(), supply_steam=steam(),
                last_effect_pressure_psia=last_psia, target_brix_out=65,
                effect_areas_ft2=[12000, 12000, 12000, 12000],
                vapor_bleeds=[0, 0, 0], dessin_coefficient=18000, liquid_level_ft=2,
            ),
            EvaporatorSet(
                juice_in=juice(), supply_steam=steam(),
                last_effect_pressure_psia=last_psia, target_brix_out=65,
                effect_areas_ft2=[11000, 9000, 9000],
                vapor_bleeds=[0, 0], dessin_coefficient=18000, liquid_level_ft=2,
            ),
            EvaporatorSet(
                juice_in=juice(), supply_steam=steam(),
                last_effect_pressure_psia=last_psia, target_brix_out=65,
                effect_areas_ft2=[15000, 15000, 15000, 15000],
                vapor_bleeds=[0, 0, 0], dessin_coefficient=18000, liquid_level_ft=2,
            ),
        ]

    def make_pre(vapor_bleed_lb_per_hr):
        """Pre evaporator (35 000 ft²) on exhaust steam. Bleeds V1 only."""
        return PreEvaporator(
            juice_in=SugarStream(
                brix=14, purity=90, flow_lb_per_hr=total_juice,
                temp_deg_F=225, pressure_psia=60, level_ft=0,
            ),
            supply_steam=EvaporatorSteam(P_psia=exh_psia, flow_lb_per_hr=0),
            vapor_bleed_lb_per_hr=vapor_bleed_lb_per_hr,
            area_ft2=_PRE_AREA,
            dessin_coefficient=18_000,
            liquid_level_ft=2,
        )

    # Bleed toggle profiles  (one list per set; length = number_of_effects - 1)
    # Set 1: 4-eff → 3 slots  |  Set 2: 4-eff → 3 slots
    # Set 3: 3-eff → 2 slots  |  Set 4: 4-eff → 3 slots
    BLEEDS_ON  = [[True,  True,  True],  [True,  True,  True],  [True,  True],  [True,  True,  True]]
    BLEEDS_OFF = [[False, False, False], [False, False, False], [False, False], [False, False, False]]
    # MIX: all sets share V1; only S1 and S4 (big surface) supply V2.
    BLEEDS_MIX = [[True,  True,  False], [True,  False, False], [True,  False],  [True,  True,  False]]

    # ---------------------------------------------------------------------------
    # USER-SPECIFIED BLEED FRACTIONS
    # ---------------------------------------------------------------------------
    # Each base scenario has two optional fraction dicts:
    #   fracs_on  — used when Pre is ON   (include 'pre' key for Pre's share)
    #   fracs_off — used when Pre is OFF  (set keys only)
    #
    # Format:  {effect_number: {participant: fraction, ...}, ...}
    #   effect_number : 1 = V1, 2 = V2, 3 = V3, ...
    #   participant   : 'pre' for the Pre evaporator (preON only)
    #                   0 = S1, 1 = S2, 2 = S3, 3 = S4  (set indices)
    #   fraction      : share of the total demand for that effect (0.0–1.0)
    #
    # Rules:
    #   • Fractions for all participants in an effect should sum to 1.0.
    #   • Only specify fractions for ONLINE sets (offline sets → omit).
    #   • Any effect not listed uses area-weighted distribution (default).
    #   • Set None for either dict to use area-weighted for all effects.
    #
    # Example — 2 sets (S1+S2) online, Pre ON, bleeds ON:
    #   fracs_on  = {1: {'pre': 0.20, 0: 0.50, 1: 0.30},   # V1 split
    #                2: {0: 0.60, 1: 0.40}}                  # V2 split
    #   fracs_off = {1: {0: 0.60, 1: 0.40},                 # V1 split (no pre)
    #                2: {0: 0.60, 1: 0.40}}                  # V2 split
    # ---------------------------------------------------------------------------
    fracs_on_3sets = {1: {'pre': 0.9, 0: 0.04, 1: 0.03, 2: 0.03}, 2: {0: 0.6, 1: 0.25, 2: 0.25}} # V1 and V2 split respectively
    fracs_off_3sets = {1: {0: 0.6, 1: 0.25, 2: 0.25}, 2: {0: 0.6, 1: 0.25, 2: 0.25}} # for no pre
    fracs_on_4sets = {1: {'pre': 0.9, 0: 0.03, 1: 0.02, 2: 0.02, 3: .03}, 2: {0: 0.5, 1: 0.2, 2: 0.2, 3: 0.1}} # V1 and V2 split respectively
    # Base scenarios: (online mask, bleed toggles, label, fracs_on, fracs_off).
    # Each base is expanded into Pre ON and Pre OFF variants, both get Excel files.
    # Set fracs_on / fracs_off to None to use area-weighted distribution (default).
    _BASE = [
        # ── 1 set online ──────────────────────────────────────────────────
        ([True,  False, False, False], BLEEDS_ON,  "1-set  S1 only          bleeds ON",  None, None),
        ([False, True,  False, False], BLEEDS_ON,  "1-set  S2 only          bleeds ON",  None, None),
        ([False, False, True,  False], BLEEDS_ON,  "1-set  S3 only          bleeds ON",  None, None),
        ([False, False, False, True ], BLEEDS_ON,  "1-set  S4 only          bleeds ON",  None, None),
        ([True,  False, False, False], BLEEDS_OFF, "1-set  S1 only          bleeds OFF", None, None),
        # ── 2 sets online ─────────────────────────────────────────────────
        ([True,  True,  False, False], BLEEDS_ON,  "2-sets S1+S2            bleeds ON",  {1: {'pre': 0.90, 0: 0.070, 1: 0.030}, 2: {0: 0.60, 1: 0.40}}, {1: {0: 0.60, 1: 0.40}, 2: {0: 0.60, 1: 0.40}} ),
        ([True,  False, True,  False], BLEEDS_ON,  "2-sets S1+S3 (diff eff) bleeds ON",  None, None),
        ([True,  False, False, True ], BLEEDS_ON,  "2-sets S1+S4            bleeds ON",  None, None),
        ([False, True,  True,  False], BLEEDS_ON,  "2-sets S2+S3            bleeds ON",  None, None),
        ([False, False, True,  True ], BLEEDS_ON,  "2-sets S3+S4 (diff eff) bleeds ON",  None, None),
        ([True,  True,  False, False], BLEEDS_OFF, "2-sets S1+S2            bleeds OFF", None, None),
        ([True,  True,  False, False], BLEEDS_MIX, "2-sets S1+S2            bleeds MIX", None, None),
        # ── 3 sets online ─────────────────────────────────────────────────
        ([True,  True,  True,  False], BLEEDS_ON,  "3-sets S1+S2+S3         bleeds ON",  fracs_on_3sets, None),
        ([True,  True,  False, True ], BLEEDS_ON,  "3-sets S1+S2+S4         bleeds ON",  None, None),
        ([True,  False, True,  True ], BLEEDS_ON,  "3-sets S1+S3+S4         bleeds ON",  None, None),
        ([False, True,  True,  True ], BLEEDS_ON,  "3-sets S2+S3+S4         bleeds ON",  None, None),
        ([True,  True,  True,  False], BLEEDS_OFF, "3-sets S1+S2+S3         bleeds OFF", None, None),
        ([True,  True,  True,  False], BLEEDS_MIX, "3-sets S1+S2+S3         bleeds MIX", None, None),
        # ── 4 sets online ─────────────────────────────────────────────────
        ([True,  True,  True,  True ], BLEEDS_ON,  "4-sets all              bleeds ON",  fracs_on_4sets, None),
        ([True,  True,  True,  True ], BLEEDS_OFF, "4-sets all              bleeds OFF", None, None),
        ([True,  True,  True,  True ], BLEEDS_MIX, "4-sets all              bleeds MIX", None, None),
    ]

    # Expand each base scenario into Pre ON and Pre OFF variants; every scenario
    # gets its own Excel file, named from the label.
    scenarios = []
    for online, toggles, label, fracs_on, fracs_off in _BASE:
        for pre_on in (True, False):
            pre_tag    = "preON" if pre_on else "preOFF"
            full_label = f"[{pre_tag}] {label}"
            slug       = re.sub(r'[^a-z0-9]+', '_', label.strip().lower()).strip('_')
            xls        = os.path.join(_HERE, f"{pre_tag}_{slug}.xlsx")
            fracs      = fracs_on if pre_on else fracs_off
            scenarios.append((online, toggles, full_label, xls, pre_on, fracs))

    PASS_SPREAD = 0.08    # U-ratio spread threshold for a "pass"; 3-eff vs 4-eff
                          # sets settle at genuinely different U ratios (~0.03-0.07)

    print("=" * 72)
    print("STRESS TEST — multi_set_solver.py")
    print(f"  Total juice feed : {total_juice:,.0f} lb/hr")
    print(f"  V1 demand        : {demands[1]:,.0f} lb/hr")
    print(f"  V2 demand        : {demands[2]:,.0f} lb/hr")
    print("=" * 72)

    results = []
    for online, toggles, label, xls, pre_on, fracs in scenarios:

        # ── Pre evaporator ────────────────────────────────────────────────
        if pre_on:
            # Use user-specified Pre V1 fraction if provided, else area-weighted.
            if fracs and 1 in fracs and 'pre' in fracs[1]:
                pre_v1_bleed = demands.get(1, 0) * fracs[1]['pre']
            else:
                v1_hs_sets = sum(
                    _SET_V1_AREAS[i]
                    for i, on in enumerate(online)
                    if on and i < len(toggles) and toggles[i] and toggles[i][0]
                )
                total_v1_hs  = _PRE_AREA + v1_hs_sets
                pre_v1_bleed = demands.get(1, 0) * _PRE_AREA / total_v1_hs if total_v1_hs > 0 else 0.0

            pre           = make_pre(pre_v1_bleed)
            pre_out       = pre.juice_out
            juice_to_sets = pre_out.flow_lb_per_hr
            sets = make_sets(
                juice_brix        = pre_out.brix,
                juice_temp_deg_F  = pre_out.temp_deg_F,
                juice_flow_per_set= juice_to_sets / len(NAMES),
            )
            remaining_demands = dict(demands)
            if 1 in remaining_demands:
                remaining_demands[1] = max(0.0, remaining_demands[1] - pre_v1_bleed)
        else:
            pre               = None
            juice_to_sets     = total_juice
            sets              = make_sets()
            remaining_demands = demands

        # ── Convert user fractions → bleed_overrides (lb/hr per set) ─────
        # Fractions are expressed as shares of the TOTAL demand for each effect.
        # The 'pre' key (V1 only) is handled above; here we build set overrides.
        bleed_overrides = None
        if fracs:
            overrides = {}
            for effect_num, effect_fracs in fracs.items():
                total_demand = demands.get(effect_num, 0)
                set_alloc = {
                    participant: frac * total_demand
                    for participant, frac in effect_fracs.items()
                    if participant != 'pre'
                }
                if set_alloc:
                    overrides[effect_num] = set_alloc
            if overrides:
                bleed_overrides = overrides

        t0 = time.time()

        solve_multi_set(
            evaporator_sets=sets,
            online=online,
            total_juice_flow_lb_per_hr=juice_to_sets,
            vapor_demands=remaining_demands,
            bleed_toggles=toggles,
            set_names=NAMES,
            excel_path=xls,
            scenario_label=label.strip(),
            verbose=False,
            pre_evaporator=pre,
            bleed_overrides=bleed_overrides,
        )

        elapsed = time.time() - t0

        active = [i for i, o in enumerate(online) if o]
        u_vals = [_mean_u_ratio(sets[i]) for i in active]
        finite = all(_is_finite(u) for u in u_vals)
        spread = (max(u_vals) - min(u_vals)) if (finite and u_vals) else float("nan")
        passed = finite and (spread <= PASS_SPREAD or len(active) == 1)
        status = "PASS" if passed else "FAIL"
        results.append((status, label, elapsed, spread, u_vals))

        u_str = "  ".join(f"S{active[j]+1}:{u:.4f}" for j, u in enumerate(u_vals))
        print(f"  [{status}]  {label:<52}  spread={spread:.4f}  {elapsed:.2f}s  |  {u_str}")
        if xls:
            print(f"           Excel -> {xls}")

    print()
    print("=" * 72)
    fails  = [r for r in results if r[0] == "FAIL"]
    passes = len(results) - len(fails)
    print(f"  Results : {passes}/{len(results)} passed")
    if fails:
        print("  FAILURES:")
        for r in fails:
            print(f"    {r[1]}  spread={r[3]:.4f}  U={r[4]}")
    print("=" * 72)
